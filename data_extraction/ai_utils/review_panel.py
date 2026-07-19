"""
Reusable AI-review workbench (mixin) shared by the review UIs.

`AIReviewMixin` provides the complete AI Review page: a shared sections
queue, the free-form review mode (presets, answer-format directive,
pre-run storage dialog with per-result auto-save, export) and the
column-analysis mode (column definitions with unique-element checkboxes,
live editable prompt preview, per-section or per-column LLM calls,
pause/resume/stop, row-by-row saving into an accumulating Excel
workbook with per-run snapshot sheets and Uniq sheets via
scripts/excel_file_utils, table export), the evaluation tab (run the
selected checks — substring grounding, Jaccard, ROUGE, BLEU, Levenshtein,
similarity ratio, WER — on a stored analysis workbook against the section
texts that generated it, into the same or a new workbook; with the
auto-evaluate box ticked each section row is evaluated right after it is
saved, via data_extraction.evaluation.column_evaluator),
plus the LLM service/model picker and the API-key manager button.

Host contract — a class mixing this in must provide:
  * ``self.root``          — the Tk root/host widget (for after()/dialogs)
  * ``self.status_var``    — a tk.StringVar used for status messages
  * ``_ai_current_section()``  -> (title, text) or None
  * ``_ai_checked_sections()`` -> list of (title, text) in document order
and call ``_init_ai_state()`` in its ``__init__`` and
``_build_ai_page(parent)`` to build the page into a container widget.
Extracted from data_extraction/easa/json_review_ui.py so the EASA studio
and the chunking studio share one implementation.
"""

import json
import os
import queue
import re
import sys
import threading
import time
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, scrolledtext, ttk

from . import llm_utils as _llm_utils_mod
from .entity_chains import is_entity_column as _is_entity_column
from .llm_utils import (get_selected_model, list_available_models,
                        set_selected_model, list_embedding_models,
                        get_default_embedding_model)

# Preset review instructions (label -> prompt sent with the section text).
# "Custom…" leaves the box for the user to write their own.
_AI_PRESETS = {
    "Summarize this section": "Summarize this regulatory section concisely.",
    "List key obligations / requirements":
        "List the key obligations or requirements in this section as short bullet points.",
    "Explain in plain language": "Explain this section in plain, non-legal language.",
    "Extract defined terms": "Extract every defined term in this section and its definition.",
    "Find references to other rules":
        "List any references this section makes to other regulations, articles or sections.",
    "Custom…": "",
}

# UI service label -> llm_call service code
_AI_SERVICES = {"Blablador": "b", "DLR Ollama": "o", "Local model": "l"}

# Answer-format label -> directive appended to the LLM prompt ("" = free-form)
_AI_OUTPUT_FORMATS = {
    "Plain text": "",
    "Markdown": "Format your entire answer as Markdown.",
    "JSON": ("Return your entire answer as a single valid JSON object — "
             "no code fences, no commentary outside the JSON."),
    "CSV table": ("Return your entire answer as CSV with a header row — "
                  "no code fences, no commentary outside the CSV."),
}


# Evaluation tab options: UI label -> column_evaluator metric name(s).
_EVAL_METRIC_OPTIONS = [
    ("Substring check (item grounded in reference)", ("grounding",)),
    ("Jaccard similarity", ("jaccard",)),
    ("ROUGE-1/2/L", ("rouge1", "rouge2", "rougeL")),
    ("BLEU", ("bleu",)),
    ("Levenshtein distance", ("levenshtein_distance",)),
    ("Similarity ratio", ("similarity_ratio",)),
    ("Word error rate (WER)", ("word_error_rate",)),
    ("Embedding cosine similarity (semantic)", ("embedding_cosine",)),
    ("BERTScore (semantic P/R/F1)",
     ("bertscore_p", "bertscore_r", "bertscore_f1")),
]

# Hover explanations for the Evaluation tab's metric checkboxes.
_EVAL_METRIC_TIPS = {
    "Substring check (item grounded in reference)":
        "Splits each cell into its ';'-items and checks every item as a "
        "case-insensitive substring of the reference text. Gives True/False "
        "counts per column and a per-section Coverage % — higher is better.",
    "Jaccard similarity":
        "Overlap of the word sets of cell and reference: |common| / |union|. "
        "0–1, higher is better. Needs no extra library.",
    "ROUGE-1/2/L":
        "Recall-oriented n-gram overlap: ROUGE-1 single words, ROUGE-2 word "
        "pairs, ROUGE-L longest common subsequence. 0–1, higher is better.",
    "BLEU":
        "N-gram precision with a brevity penalty (machine-translation "
        "metric). 0–1, higher is better; very short texts often score low.",
    "Levenshtein distance":
        "Number of single-character edits (insert/delete/replace) to turn "
        "the reference into the cell text. 0+, LOWER is better.",
    "Similarity ratio":
        "Normalised edit similarity of the two texts. 0–1, higher is "
        "better (1 = identical).",
    "Word error rate (WER)":
        "Word-level edits divided by the reference length (speech-"
        "recognition metric). 0+, LOWER is better; can exceed 1.",
    "Embedding cosine similarity (semantic)":
        "Cosine similarity of the two texts' embedding vectors — measures "
        "MEANING, not wording. −1–1, higher is better. Uses the embedding "
        "backend configured below.",
    "BERTScore (semantic P/R/F1)":
        "Token-level semantic precision/recall/F1 from per-token embedding "
        "matches. Higher is better. Local embedding backend only (the "
        "remote API returns one pooled vector per string).",
    "__uniq__":
        "Additionally collects the unique generated values per column "
        "(Uniq sheet) and grounds each against the combined reference "
        "text (Uniq Eval sheet with coverage summaries).",
}


class _Tooltip:
    """Small hover tooltip for a widget (delay, wraps long text)."""

    def __init__(self, widget, text, delay=500):
        self.widget, self.text, self.delay = widget, text, delay
        self._after = None
        self._tip = None
        widget.bind("<Enter>", self._schedule, add="+")
        widget.bind("<Leave>", self._hide, add="+")
        widget.bind("<ButtonPress>", self._hide, add="+")

    def _schedule(self, _event=None):
        self._cancel()
        self._after = self.widget.after(self.delay, self._show)

    def _cancel(self):
        if self._after:
            self.widget.after_cancel(self._after)
            self._after = None

    def _show(self):
        if self._tip:
            return
        x = self.widget.winfo_rootx() + 12
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 4
        self._tip = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        tk.Label(tw, text=self.text, wraplength=380, justify="left",
                 background="#ffffe0", relief="solid", borderwidth=1,
                 padx=6, pady=4).pack()

    def _hide(self, _event=None):
        self._cancel()
        if self._tip:
            self._tip.destroy()
            self._tip = None


# BERTScore is local-backend-only: it needs one vector per TOKEN, and the remote
# /embeddings endpoint returns a single pooled vector per string. The tab
# disables this option on the API backend rather than let it silently duplicate
# embedding_cosine (see _eval_sync_bertscore_state).
_BERTSCORE_LABEL = "BERTScore (semantic P/R/F1)"

# Uniques tab: UI label -> item separator inside one cell. ``None`` auto-detects
# the most frequent among the common candidates; ``"custom"`` reads the free
# text field so the user can type any separator (e.g. " / ", " and ").
_UNIQ_SEPARATORS = [
    ("Auto-detect (most frequent)", None),
    ("Semicolon   ;", ";"),
    ("Comma   ,", ","),
    ("Pipe   |", "|"),
    ("Newline", "\n"),
    ("Tab", "\t"),
    ("Custom…", "custom"),
]
# Candidates tried when auto-detecting a cell's separator.
_UNIQ_AUTO_CANDIDATES = (";", "|", "\n", "\t", ",")

# The column definitions of the Column-analysis tab persist here (same config
# folder as the API keys) so they survive closing and reopening the studios.
def _col_config_path():
    from .file_manager import ALR_main_folder
    return Path(ALR_main_folder) / "column_analysis_columns.json"


# Session state of the AI Review page (storage file, service/model, evaluation
# choices, window geometry/sash) — restored on the next launch.
def _ui_state_path():
    from .file_manager import ALR_main_folder
    return Path(ALR_main_folder) / "review_ui_state.json"


# Extra attempts when an LLM reply is not parsable JSON: the same prompt is
# re-sent up to this many times before the value is marked [unparsed].
_JSON_RETRIES = 3

# Placeholder appended to the editable prompt preview; everything above the
# "---" line is the prompt head, the placeholder is replaced by the real
# section at run time.
_PREVIEW_TAIL = "\n\n---\nSECTION: <title>\n\n<section text>"

_EXCEL_UTILS = None


def _excel_utils():
    """Load scripts/excel_file_utils.py (repo root) once, by file path."""
    global _EXCEL_UTILS
    if _EXCEL_UTILS is None:
        import importlib.util
        p = Path(__file__).resolve().parents[2] / "scripts" / "excel_file_utils.py"
        spec = importlib.util.spec_from_file_location("excel_file_utils", str(p))
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        _EXCEL_UTILS = mod
    return _EXCEL_UTILS


class AIReviewMixin:
    """Complete AI Review page; see module docstring for the host contract."""

    def _init_ai_state(self):
        self._ai_batch = []           # queued (title, text) sections for looped AI runs
        self._ai_results = []         # accumulated free-form AI results for export
        self._ai_busy = False
        self._ai_columns = [          # (name, what the LLM should extract)
            ("Summary", "A one-sentence summary of the section."),
            ("References", "Other rules, articles or sections this section refers to."),
        ]
        self._col_results = []        # column-analysis rows (dicts) for export
        self._col_table_cols = []     # column names currently shown in the table
        self._col_uniq_checked = {name for name, _ in self._ai_columns}
        self._col_analyze_checked = {name for name, _ in self._ai_columns}
        self._col_load_columns()      # restore the previous session's columns
        self._col_run_uniq_cols = []  # checked columns snapshotted at run start
        self._ai_autosave_path = None # where the current free-form run auto-saves
        self._ai_run_start = 0        # index of the first result of the current run
        self._ai_save_error_shown = False
        self._col_store_path = None   # workbook the column analyses accumulate into
        self._col_rows_saved = 0      # rows of the current run already on disk
        self._col_run_sheet = None    # snapshot sheet of the current run (.xlsx)
        self._col_save_error_shown = False
        self._col_run_refs = {}       # {section title: text} of the current run
        self._col_run_auto_eval = False  # evaluate each row as it is saved
        self._col_eval_choices = {}   # picker state (label -> bool, "__uniq__")
        self._col_run_eval_metrics = []  # evaluations chosen for this run
        self._col_run_eval_uniq = False  # also evaluate unique values per row
        self._col_eval_entries = []   # per-section evaluations of this run
        self._col_pause_ev = threading.Event()  # set = worker pauses
        self._col_stop_ev = threading.Event()   # set = worker stops ASAP
        self._run_t0 = None           # wall-clock start of the current run
        self._ai_call_count = 0       # LLM calls (incl. retries) of the run
        self._run_current = 0         # 1-based index of the section in work
        self._run_queue_iids = []     # queue rows snapshotted at run start
        self._col_rerun = False       # current run re-does failed rows only
        self._col_run_ctx = {}        # settings snapshot for "Re-run failed"
        self._ai_last_output = None   # latest result file (Open results)
        self._eval_busy = False       # an evaluation is running
        self._uniq_busy = False       # a manual unique-elements run is in progress
        self._uniq_col_vars = {}      # {column name: BooleanVar} for the Uniques tab

    # -- host hooks --------------------------------------------------------- #
    def _ai_current_section(self):
        """Return (title, text) for the host's current section, or None."""
        raise NotImplementedError

    def _ai_checked_sections(self):
        """Return the host's checked sections as [(title, text), ...]."""
        raise NotImplementedError

    def _default_export_name(self, kind, ext):
        stem = Path(getattr(self, "json_path", None) or "review").stem
        return f"{stem}_{kind}.{ext}"

    def _build_ai_page(self, parent):
        # A cleaner ttk look for the whole app (no-op when unavailable).
        try:
            style = ttk.Style(self.root)
            if "clam" in style.theme_names():
                style.theme_use("clam")
        except tk.TclError:
            pass

        ai = ttk.Frame(parent, padding=6)
        ai.pack(fill="both", expand=True)

        # LLM row: service, inline model picker (↻ loads the live list in the
        # background) and a chip always showing what a run would use NOW.
        llm_frame = ttk.Frame(ai)
        llm_frame.pack(fill="x", padx=10, pady=5)

        ttk.Label(llm_frame, text="LLM Processing Service Engine:").pack(side="left", padx=5)
        self.llm_choice_an = ttk.Combobox(llm_frame, values=["O", "B"], width=5, state="readonly")
        self.llm_choice_an.set("O")
        self.llm_choice_an.pack(side="left", padx=5)
        self.llm_choice_an.bind("<<ComboboxSelected>>", self._llm_service_changed)

        ttk.Label(llm_frame, text="Model:").pack(side="left", padx=(8, 0))
        self.llm_model_combo = ttk.Combobox(llm_frame, state="readonly", width=28,
                                            values=[])
        self.llm_model_combo.pack(side="left", padx=4)
        self.llm_model_combo.bind("<<ComboboxSelected>>", self._llm_on_model_pick)
        ttk.Button(llm_frame, text="↻", width=3,
                   command=self._llm_models_refresh).pack(side="left")
        self.llm_active_lbl = ttk.Label(llm_frame, foreground="#1a6fb0")
        self.llm_active_lbl.pack(side="left", padx=(10, 0))

        ttk.Button(llm_frame, text="API keys…", command=self._ai_manage_keys).pack(side="right", padx=5)

        # Vertical split: analysis area on top (~80%), console pane below
        # (~20% of the window) so the log stays visible; drag to adjust.
        vsplit = ttk.PanedWindow(ai, orient="vertical")
        vsplit.pack(fill="both", expand=True, pady=(6, 0))
        self._ai_vsplit = vsplit

        body = ttk.PanedWindow(vsplit, orient="horizontal")
        vsplit.add(body, weight=4)

        # Left: the shared sections queue both run modes consume. A Treeview
        # so every section shows its live status during a run
        # (⏳ queued / ▶ running / ✔ done / ✖ error).
        batch_wrap = ttk.LabelFrame(body, text=" Sections queue ", padding=4)
        body.add(batch_wrap, weight=1)
        self.ai_batch_list = ttk.Treeview(batch_wrap, columns=("st",),
                                          show="tree headings", height=4,
                                          selectmode="extended")
        self.ai_batch_list.heading("#0", text="Section")
        self.ai_batch_list.heading("st", text="Status")
        self.ai_batch_list.column("#0", width=170, anchor="w")
        self.ai_batch_list.column("st", width=78, minwidth=64,
                                  anchor="center", stretch=False)
        for tag, color in (("pending", "grey35"), ("running", "#1a6fb0"),
                           ("done", "#1f7a1f"), ("error", "#b00020")):
            self.ai_batch_list.tag_configure(tag, foreground=color)
        self.ai_batch_list.pack(fill="both", expand=True)
        qbtns = ttk.Frame(batch_wrap)
        qbtns.pack(fill="x", pady=(4, 0))
        ttk.Button(qbtns, text="＋ Current node", command=self._ai_add_to_batch).pack(side="left")
        ttk.Button(qbtns, text="＋ Checked", command=self._ai_add_checked).pack(side="left", padx=4)
        ttk.Button(qbtns, text="Remove", command=self._ai_remove_selected).pack(side="left")
        ttk.Button(qbtns, text="Clear", command=self._ai_clear_batch).pack(side="left", padx=4)

        # Right: free-form review | column analysis
        self.ai_nb = ttk.Notebook(body)
        body.add(self.ai_nb, weight=3)
        self._build_ai_freeform_tab()
        self._build_ai_columns_tab()
        self._build_ai_eval_tab()
        self._build_ai_uniques_tab()

        # Bottom pane: progress bar + console logging every step of the runs
        # (what is analyzed / sent to the LLM / answered / evaluated).
        prog_wrap = ttk.LabelFrame(vsplit, text=" Progress ", padding=4)
        vsplit.add(prog_wrap, weight=1)
        bar_row = ttk.Frame(prog_wrap)
        bar_row.pack(fill="x")
        self.ai_progress = ttk.Progressbar(bar_row, mode="determinate")
        self.ai_progress.pack(side="left", fill="x", expand=True)
        # Unmissable pause indicator + live "done/total · ETA · calls" info.
        self.ai_paused_lbl = ttk.Label(bar_row, text="", foreground="#b8860b")
        self.ai_paused_lbl.pack(side="left", padx=(8, 0))
        self.ai_progress_info = ttk.Label(bar_row, text="", foreground="grey40")
        self.ai_progress_info.pack(side="right", padx=(8, 0))

        # Console toolbar: filter (live, hides non-matching lines),
        # auto-scroll toggle, open-last-result and save-log actions.
        tools = ttk.Frame(prog_wrap)
        tools.pack(fill="x", pady=(4, 0))
        ttk.Label(tools, text="Filter:").pack(side="left")
        self.ai_log_filter = ttk.Entry(tools, width=18)
        self.ai_log_filter.pack(side="left", padx=4)
        self.ai_log_filter.bind("<KeyRelease>", self._ai_log_refilter)
        self.ai_autoscroll_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(tools, variable=self.ai_autoscroll_var,
                        text="Auto-scroll").pack(side="left", padx=8)
        ttk.Button(tools, text="Save log…",
                   command=self._ai_log_save).pack(side="right")
        self.ai_open_btn = ttk.Button(tools, text="Open results",
                                      state="disabled",
                                      command=self._ai_open_results)
        self.ai_open_btn.pack(side="right", padx=6)

        self.ai_console = scrolledtext.ScrolledText(
            prog_wrap, height=6, wrap="word", state="disabled")
        self.ai_console.pack(fill="both", expand=True, pady=(4, 0))
        # Severity colours + an elide tag the filter uses to hide lines.
        self.ai_console.tag_configure("err", foreground="#b00020")
        self.ai_console.tag_configure("warn", foreground="#b8860b")
        self.ai_console.tag_configure("ok", foreground="#1f7a1f")
        self.ai_console.tag_configure("hdr", foreground="#1a6fb0")
        self.ai_console.tag_configure("hidden", elide=True)

        # Place the sash once — at the saved position from the last session
        # when there is one, else at ~80/20 so the console gets a fifth of
        # the window (weights only govern later resizes).
        def _place_sash(_event=None):
            height = vsplit.winfo_height()
            if height > 120 and not getattr(vsplit, "_sash_placed", False):
                vsplit._sash_placed = True
                pos = getattr(self, "_ui_saved_sash", None)
                if not isinstance(pos, int) or not 60 <= pos <= height - 60:
                    pos = int(height * 0.8)
                try:
                    vsplit.sashpos(0, pos)
                except tk.TclError:
                    pass
        vsplit.bind("<Configure>", _place_sash)

        # Restore last session's choices (built last: all widgets exist now).
        self._ui_state_load()
        self._llm_update_chip()

        # Keyboard shortcuts: Ctrl/Cmd+Return = Analyze queued,
        # Ctrl/Cmd+P = Pause/Resume, Ctrl/Cmd+. = Stop.
        top = self.root.winfo_toplevel()
        for seq in ("<Control-Return>", "<Command-Return>"):
            top.bind(seq, self._kb_run, add="+")
        for seq in ("<Control-p>", "<Command-p>"):
            top.bind(seq, self._kb_pause, add="+")
        for seq in ("<Control-period>", "<Command-period>"):
            top.bind(seq, self._kb_stop, add="+")

    # ------------------------------------------------- keyboard shortcuts -- #
    def _kb_run(self, _event=None):
        if not self._ai_busy and not self._eval_busy:
            self._col_run()
        return "break"

    def _kb_pause(self, _event=None):
        if self._ai_busy and str(self.col_pause_btn.cget("state")) != "disabled":
            self._col_pause_resume()
        return "break"

    def _kb_stop(self, _event=None):
        if self._ai_busy and str(self.col_stop_btn.cget("state")) != "disabled":
            self._col_stop()
        return "break"

    # -------------------------------------------- inline LLM model picker -- #
    def _llm_service_label(self):
        return "DLR Ollama" if self.llm_choice_an.get().upper() == "O" else "BlaBla"

    def _llm_update_chip(self):
        """Show what a run started NOW would use (service + model)."""
        label = self._llm_service_label()
        model = get_selected_model(label)
        self.llm_active_lbl.configure(text=f"Active: {label} · {model or 'default'}")
        current = model or ""
        if current and current not in (self.llm_model_combo.cget("values") or ()):
            self.llm_model_combo.configure(
                values=[current, *self.llm_model_combo.cget("values")])
        self.llm_model_combo.set(current)

    def _llm_service_changed(self, _event=None):
        self.llm_model_combo.configure(values=[])
        self._llm_update_chip()
        self._ui_state_save()

    def _llm_on_model_pick(self, _event=None):
        model = self.llm_model_combo.get().strip()
        if model:
            set_selected_model(self._llm_service_label(), model)
        self._llm_update_chip()
        self._ui_state_save()

    def _llm_models_refresh(self):
        """Load the service's live model list in the background and fill the
        inline picker (the UI stays responsive; errors go to the status bar)."""
        label = self._llm_service_label()
        self.status_var.set(f"Loading {label} models…")

        def work():
            try:
                models = list_available_models(label) or []
            except Exception as exc:  # noqa: BLE001
                models, err = [], str(exc)
            else:
                err = None

            def apply():
                if err or not models:
                    self.status_var.set(
                        f"Could not list {label} models"
                        + (f": {err}" if err else " (empty list)."))
                    return
                self.llm_model_combo.configure(values=models)
                current = get_selected_model(label)
                if current in models:
                    self.llm_model_combo.set(current)
                self.status_var.set(f"{len(models)} {label} model(s) loaded.")
            try:
                self.root.after(0, apply)
            except (RuntimeError, tk.TclError):
                pass  # app closing / no mainloop

        threading.Thread(target=work, daemon=True).start()

    # -------------------------------------------------- progress & console -- #
    @staticmethod
    def _ai_log_severity(msg):
        """Colour tag for a console line, by content."""
        if "[ERROR]" in msg or "❌" in msg:
            return "err"
        if "⚠" in msg or "↻" in msg or "[WARN]" in msg or "[unparsed]" in msg:
            return "warn"
        if msg.startswith("===") or msg.startswith("⚖"):
            return "hdr"
        if "💾" in msg or "✅" in msg or "✔" in msg or "→ wrote" in msg:
            return "ok"
        return None

    def _ai_log(self, msg):
        """Append one line to the progress console (main thread only):
        severity-coloured, hidden immediately when it doesn't match the
        active filter, auto-scrolled only while the toggle is on."""
        from datetime import datetime
        line = f"[{datetime.now().strftime('%H:%M:%S')}] {msg}\n"
        tags = []
        sev = self._ai_log_severity(str(msg))
        if sev:
            tags.append(sev)
        filt = self.ai_log_filter.get().strip().lower()
        if filt and filt not in line.lower():
            tags.append("hidden")
        self.ai_console.configure(state="normal")
        self.ai_console.insert("end", line, tuple(tags))
        # keep the console bounded
        if int(self.ai_console.index("end-1c").split(".")[0]) > 2000:
            self.ai_console.delete("1.0", "500.0")
        if self.ai_autoscroll_var.get():
            self.ai_console.see("end")
        self.ai_console.configure(state="disabled")

    def _ai_log_refilter(self, _event=None):
        """Re-apply the console filter to every existing line (elide
        non-matching ones; an empty filter shows everything again)."""
        filt = self.ai_log_filter.get().strip().lower()
        self.ai_console.configure(state="normal")
        self.ai_console.tag_remove("hidden", "1.0", "end")
        if filt:
            last = int(self.ai_console.index("end-1c").split(".")[0])
            for ln in range(1, last + 1):
                text = self.ai_console.get(f"{ln}.0", f"{ln}.end")
                if text and filt not in text.lower():
                    self.ai_console.tag_add("hidden", f"{ln}.0", f"{ln + 1}.0")
        if self.ai_autoscroll_var.get():
            self.ai_console.see("end")
        self.ai_console.configure(state="disabled")

    def _ai_log_save(self):
        """Write the whole console log (unfiltered) to a text file."""
        from datetime import datetime
        path = filedialog.asksaveasfilename(
            title="Save the progress log as…",
            defaultextension=".log",
            filetypes=[("Log", "*.log"), ("Text", "*.txt")],
            initialfile=f"ai_review_{datetime.now().strftime('%Y-%m-%d_%H%M')}.log")
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8") as f:
                f.write(self.ai_console.get("1.0", "end-1c"))
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("Save log", str(exc))
            return
        self.status_var.set(f"Progress log saved → {path}")

    def _ai_note_output(self, path):
        """Remember the latest result file and arm the Open-results button."""
        if path:
            self._ai_last_output = str(path)
            self.ai_open_btn.configure(state="normal")

    def _ai_open_results(self):
        """Open the latest result file with the system's default app."""
        path = getattr(self, "_ai_last_output", None)
        if not path or not os.path.exists(path):
            self.status_var.set("No result file to open yet.")
            return
        try:
            if sys.platform == "darwin":
                import subprocess
                subprocess.Popen(["open", path])
            elif os.name == "nt":
                os.startfile(path)  # noqa: S606 - user-chosen file
            else:
                import subprocess
                subprocess.Popen(["xdg-open", path])
            self.status_var.set(f"Opened {os.path.basename(path)}.")
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("Open results", str(exc))

    def _ai_log_bg(self, msg):
        """Thread-safe console logging for the worker threads."""
        self.root.after(0, self._ai_log, msg)

    @staticmethod
    def _ai_preview(text, limit=220):
        text = " ".join(str(text if text is not None else "").split())
        return text if len(text) <= limit else text[:limit] + " …"

    def _ai_progress_start(self, total):
        self.ai_progress.stop()
        self.ai_progress.configure(mode="determinate", maximum=max(total, 1), value=0)
        self._run_t0 = time.time()
        self._ai_call_count = 0
        self._run_current = 0
        self._ai_progress_info_update()

    def _ai_progress_step(self):
        self.ai_progress.configure(
            value=min(float(self.ai_progress["value"]) + 1,
                      float(self.ai_progress["maximum"])))
        self._ai_progress_info_update()

    def _ai_progress_done(self):
        self.ai_progress.stop()
        self.ai_progress.configure(mode="determinate",
                                   value=self.ai_progress["maximum"])
        self.ai_paused_lbl.configure(text="")
        self._ai_progress_info_update()

    @staticmethod
    def _fmt_eta(seconds):
        seconds = int(seconds)
        if seconds < 60:
            return f"{seconds}s"
        m, s = divmod(seconds, 60)
        if m < 60:
            return f"{m}m {s:02d}s"
        h, m = divmod(m, 60)
        return f"{h}h {m:02d}m"

    def _ai_progress_info_update(self):
        """Refresh the 'done/total · ETA · calls' label beside the bar."""
        if not self._run_t0:
            self.ai_progress_info.configure(text="")
            return
        try:
            total = int(float(self.ai_progress["maximum"]))
            done = int(float(self.ai_progress["value"]))
        except (tk.TclError, ValueError):
            return
        parts = [f"{done}/{total} section(s)"]
        if 0 < done < total:
            remaining = (time.time() - self._run_t0) / done * (total - done)
            parts.append(f"~{self._fmt_eta(remaining)} left")
        parts.append(f"{self._ai_call_count} LLM call(s)")
        self.ai_progress_info.configure(text=" · ".join(parts))

    # ------------------------------------------------ queue status column -- #
    _QUEUE_GLYPHS = {"pending": "⏳ queued", "running": "▶ running",
                     "done": "✔ done", "error": "✖ error"}

    def _queue_add(self, title):
        """Append one section to the queue Treeview (status: queued)."""
        self.ai_batch_list.insert("", "end", text=title,
                                  values=(self._QUEUE_GLYPHS["pending"],),
                                  tags=("pending",))

    def _queue_run_snapshot(self):
        """Freeze the queue rows of the starting run (so later removals can't
        shift the status mapping) and reset them all to 'queued'."""
        self._run_queue_iids = list(self.ai_batch_list.get_children())
        for iid in self._run_queue_iids:
            self.ai_batch_list.item(iid, values=(self._QUEUE_GLYPHS["pending"],),
                                    tags=("pending",))

    def _queue_snapshot_for(self, titles):
        """Map a re-run's jobs onto the queue rows with those titles (first
        match wins per title; a title no longer queued maps to None)."""
        by_text = {}
        for iid in self.ai_batch_list.get_children():
            by_text.setdefault(self.ai_batch_list.item(iid, "text"), []).append(iid)
        self._run_queue_iids = []
        for t in titles:
            matches = by_text.get(t) or []
            iid = matches.pop(0) if matches else None
            self._run_queue_iids.append(iid)
            if iid:
                self.ai_batch_list.item(
                    iid, values=(self._QUEUE_GLYPHS["pending"],),
                    tags=("pending",))

    def _queue_set_status(self, idx, status):
        """Set the status glyph of the run's idx-th (0-based) section."""
        iids = self._run_queue_iids
        if 0 <= idx < len(iids) and iids[idx] \
                and self.ai_batch_list.exists(iids[idx]):
            self.ai_batch_list.item(iids[idx],
                                    values=(self._QUEUE_GLYPHS[status],),
                                    tags=(status,))
            self.ai_batch_list.see(iids[idx])

    def _scrollable_tab(self, title):
        """Add a notebook tab whose whole content scrolls vertically, so every
        widget stays reachable no matter how small the window is. Returns the
        interior frame the tab's widgets should be packed into. When the
        window is taller than the content, the interior stretches to fill it
        (so expanding widgets keep working); when it is smaller, the tab
        scrolls — via the scrollbar or the mouse wheel (widgets that scroll
        themselves, like text boxes and tables, keep their own wheel)."""
        outer = ttk.Frame(self.ai_nb)
        self.ai_nb.add(outer, text=title)
        canvas = tk.Canvas(outer, highlightthickness=0, borderwidth=0)
        vbar = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vbar.set)
        vbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        inner = ttk.Frame(canvas, padding=6)
        win = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _layout(_event=None):
            cw = max(canvas.winfo_width(), 1)
            height = max(inner.winfo_reqheight(), canvas.winfo_height())
            # only touch the canvas item when something changed, otherwise
            # the <Configure> events re-trigger each other forever
            if (canvas.itemcget(win, "width") != str(cw)
                    or canvas.itemcget(win, "height") != str(height)):
                canvas.itemconfigure(win, width=cw, height=height)
            canvas.configure(scrollregion=(0, 0, cw, height))

        inner.bind("<Configure>", _layout)
        canvas.bind("<Configure>", _layout)

        def _wheel(event):
            w = event.widget
            if w is not canvas and isinstance(
                    w, (tk.Text, tk.Listbox, ttk.Treeview, tk.Canvas)):
                return  # that widget scrolls itself
            num = getattr(event, "num", None)
            if num == 4:            # X11 wheel up
                step = -1
            elif num == 5:          # X11 wheel down
                step = 1
            else:
                delta = getattr(event, "delta", 0)
                if not delta:
                    return
                # Windows reports multiples of 120, macOS small counts
                step = (-int(delta / 120) if abs(delta) >= 120
                        else (-1 if delta > 0 else 1))
            canvas.yview_scroll(step, "units")

        def _bind_wheel(_e):
            canvas.bind_all("<MouseWheel>", _wheel)
            canvas.bind_all("<Button-4>", _wheel)
            canvas.bind_all("<Button-5>", _wheel)

        def _unbind_wheel(_e):
            canvas.unbind_all("<MouseWheel>")
            canvas.unbind_all("<Button-4>")
            canvas.unbind_all("<Button-5>")

        canvas.bind("<Enter>", _bind_wheel)
        canvas.bind("<Leave>", _unbind_wheel)
        return inner

    def _build_ai_freeform_tab(self):
        ff = self._scrollable_tab("Free-form review")

        cfg = ttk.Frame(ff)
        cfg.pack(fill="x")
        ttk.Label(cfg, text="Preset:").pack(side="left")
        self.ai_preset = ttk.Combobox(cfg, state="readonly", width=28,
                                      values=list(_AI_PRESETS.keys()))
        self.ai_preset.pack(side="left", padx=4)
        self.ai_preset.bind("<<ComboboxSelected>>", self._ai_on_preset)
        ttk.Label(cfg, text="Answer as:").pack(side="left", padx=(10, 0))
        self.ai_format = ttk.Combobox(cfg, state="readonly", width=10,
                                      values=list(_AI_OUTPUT_FORMATS.keys()))
        self.ai_format.set("Plain text")
        self.ai_format.pack(side="left", padx=4)

        ttk.Label(ff, text="Instruction (the section's text is appended automatically):").pack(
            anchor="w", pady=(6, 0))
        self.ai_prompt = scrolledtext.ScrolledText(ff, height=2, wrap="word")
        self.ai_prompt.pack(fill="x")
        self.ai_prompt.insert("1.0", _AI_PRESETS["Summarize this section"])
        self.ai_preset.set("Summarize this section")

        btns = ttk.Frame(ff)
        btns.pack(fill="x", pady=6)
        self.ai_run_btn = ttk.Button(btns, text="Run on current node", command=self._ai_run_current)
        self.ai_run_btn.pack(side="left")
        self.ai_batch_btn = ttk.Button(btns, text="Run queued (0)", command=self._ai_run_batch)
        self.ai_batch_btn.pack(side="left", padx=4)
        ttk.Button(btns, text="Export results…", command=self._ai_export).pack(side="right")

        res_wrap = ttk.LabelFrame(ff, text=" Results ", padding=4)
        res_wrap.pack(fill="both", expand=True)
        self.ai_results = scrolledtext.ScrolledText(res_wrap, wrap="word", state="disabled")
        self.ai_results.pack(fill="both", expand=True)

    def _build_ai_columns_tab(self):
        ca = self._scrollable_tab("Column analysis")

        top = ttk.PanedWindow(ca, orient="horizontal")
        top.pack(fill="x")

        # Column definitions: every change here rebuilds the prompt preview.
        cols_wrap = ttk.LabelFrame(
            top, text=" Columns to extract (Run ✓ = analyze this column · "
                      "Uniq ✓ = collect unique elements) ", padding=4)
        top.add(cols_wrap, weight=1)
        self.col_defs = ttk.Treeview(cols_wrap, columns=("an", "uniq", "what"),
                                     show="tree headings", height=3)
        self.col_defs.heading("#0", text="Column")
        self.col_defs.heading("an", text="Run ✓",
                              command=self._col_toggle_analyze_all)
        self.col_defs.heading("uniq", text="Uniq ✓",
                              command=self._col_toggle_uniq_all)
        self.col_defs.heading("what", text="What should the LLM extract?")
        self.col_defs.column("#0", width=140, anchor="w")
        self.col_defs.column("an", width=48, minwidth=40, anchor="center", stretch=False)
        self.col_defs.column("uniq", width=48, minwidth=40, anchor="center", stretch=False)
        self.col_defs.column("what", width=330, anchor="w")
        self.col_defs.pack(fill="x")
        self.col_defs.bind("<<TreeviewSelect>>", self._col_on_select)
        self.col_defs.bind("<Button-1>", self._col_defs_click)

        edit = ttk.Frame(cols_wrap)
        edit.pack(fill="x", pady=(4, 0))
        ttk.Label(edit, text="Name:").pack(side="left")
        self.col_name = ttk.Entry(edit, width=16)
        self.col_name.pack(side="left", padx=4)
        ttk.Label(edit, text="Extract:").pack(side="left")
        self.col_what = ttk.Entry(edit)
        self.col_what.pack(side="left", fill="x", expand=True, padx=4)
        ttk.Button(edit, text="Add / Update", command=self._col_add_update).pack(side="left")
        ttk.Button(edit, text="Remove", command=self._col_remove).pack(side="left", padx=4)
        ttk.Button(edit, text="Clear all", command=self._col_clear).pack(side="left")
        ttk.Button(edit, text="＋ Specific entities",
                   command=self._col_add_entity_column).pack(side="left", padx=4)

        prev_wrap = ttk.LabelFrame(
            top, text=" Prompt preview (editable — sent per section) ", padding=4)
        top.add(prev_wrap, weight=1)
        # Editable: the text above the "---" marker is used as the prompt head
        # of the next run. Changing the column definitions rebuilds it.
        self.col_preview = scrolledtext.ScrolledText(prev_wrap, height=5, wrap="word")
        self.col_preview.pack(fill="both", expand=True)

        run_row = ttk.Frame(ca)
        run_row.pack(fill="x", pady=6)
        self.col_run_btn = ttk.Button(run_row, text="Analyze queued (0)", command=self._col_run)
        self.col_run_btn.pack(side="left")
        self.col_pause_btn = ttk.Button(run_row, text="Pause", state="disabled",
                                        command=self._col_pause_resume)
        self.col_pause_btn.pack(side="left", padx=(6, 0))
        self.col_stop_btn = ttk.Button(run_row, text="Stop", state="disabled",
                                       command=self._col_stop)
        self.col_stop_btn.pack(side="left", padx=(4, 0))
        # Enabled after a run that left [ERROR]/[unparsed] rows behind:
        # re-analyzes only those sections and updates their existing rows.
        self.col_rerun_btn = ttk.Button(run_row, text="Re-run failed",
                                        state="disabled",
                                        command=self._col_rerun_failed)
        self.col_rerun_btn.pack(side="left", padx=(10, 0))
        # When on, Analyze queued first asks which evaluations to run
        # automatically once the batch finishes (see the Evaluation tab).
        self.eval_auto_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(run_row, variable=self.eval_auto_var,
                        command=self._ui_state_save,
                        text="Auto-evaluate after the analysis").pack(side="left", padx=10)
        ttk.Button(run_row, text="Export table…", command=self._col_export).pack(side="right")

        table_wrap = ttk.LabelFrame(ca, text=" Analysis table (one row per section) ", padding=4)
        table_wrap.pack(fill="both", expand=True)
        self.col_table = ttk.Treeview(table_wrap, show="headings", height=6)
        ct_y = ttk.Scrollbar(table_wrap, orient="vertical", command=self.col_table.yview)
        ct_x = ttk.Scrollbar(table_wrap, orient="horizontal", command=self.col_table.xview)
        self.col_table.configure(yscrollcommand=ct_y.set, xscrollcommand=ct_x.set)
        ct_y.pack(side="right", fill="y")
        ct_x.pack(side="bottom", fill="x")
        self.col_table.pack(side="left", fill="both", expand=True)
        self.col_table.bind("<Double-Button-1>", self._col_show_row)

        self._col_refresh_defs()

    def _ai_manage_keys(self):
        try:
            from ..ai_utils.key_manager_ui import open_api_key_dialog
        except Exception as exc:  # pragma: no cover - ai_utils should always ship
            messagebox.showerror("API Keys", f"Key manager unavailable:\n{exc}")
            return
        open_api_key_dialog(self.root)

    def _ai_on_preset(self, event=None):
        preset = self.ai_preset.get()
        text = _AI_PRESETS.get(preset, "")
        if preset == "Custom…":
            return  # leave whatever the user typed
        self.ai_prompt.delete("1.0", tk.END)
        self.ai_prompt.insert("1.0", text)

    def _ai_append(self, s):
        self.ai_results.configure(state="normal")
        self.ai_results.insert("end", s)
        self.ai_results.see("end")
        self.ai_results.configure(state="disabled")

    def _ai_run_current(self):
        section = self._ai_current_section()
        if section is None:
            messagebox.showinfo("No section", "Select a section first.")
            return
        self._ai_run([section])

    def _ai_add_to_batch(self):
        section = self._ai_current_section()
        if section is None:
            messagebox.showinfo("No section", "Select a section first.")
            return
        self._ai_batch.append(section)
        self._queue_add(section[0])
        self._ai_update_queue_counts()

    def _ai_add_checked(self):
        sections = self._ai_checked_sections()
        if not sections:
            messagebox.showinfo(
                "Nothing checked",
                "Tick the ✓ checkboxes first (or use Select all).")
            return
        queued = {(t, txt) for t, txt in self._ai_batch}
        added = 0
        for job in sections:
            if job in queued:
                continue
            self._ai_batch.append(job)
            self._queue_add(job[0])
            added += 1
        self._ai_update_queue_counts()
        self.status_var.set(f"Added {added} checked section(s) to the AI batch"
                            + (f" ({len(sections) - added} already queued)." if added < len(sections) else "."))

    def _ai_remove_selected(self):
        items = list(self.ai_batch_list.get_children())
        picked = sorted((items.index(iid) for iid in self.ai_batch_list.selection()),
                        reverse=True)
        for i in picked:
            del self._ai_batch[i]
            self.ai_batch_list.delete(items[i])
        self._ai_update_queue_counts()

    def _ai_clear_batch(self):
        self._ai_batch = []
        self.ai_batch_list.delete(*self.ai_batch_list.get_children())
        self._ai_update_queue_counts()

    def _ai_update_queue_counts(self):
        n = len(self._ai_batch)
        self.ai_batch_btn.configure(text=f"Run queued ({n})")
        self.col_run_btn.configure(text=f"Analyze queued ({n})")

    def _ai_run_batch(self):
        if not self._ai_batch:
            messagebox.showinfo("Empty batch", "Add one or more sections to the batch first.")
            return
        fmt = self._ai_ask_format()
        if fmt is None:  # user cancelled
            return
        self.ai_format.set(fmt)
        self._ai_run(list(self._ai_batch))

    def _ai_ask_format(self):
        """Modal dialog: how should the LLM format the results? None = cancel."""
        win = tk.Toplevel(self.root)
        win.title("Result format")
        win.transient(self.root.winfo_toplevel())
        win.resizable(False, False)
        frm = ttk.Frame(win, padding=12)
        frm.pack(fill="both", expand=True)
        ttk.Label(frm, text="In which format should the LLM return the batch results?\n"
                            "(The instruction is sent to the model with every section.)").pack(anchor="w")
        var = tk.StringVar(value=self.ai_format.get() or "Plain text")
        for fmt in _AI_OUTPUT_FORMATS:
            ttk.Radiobutton(frm, text=fmt, value=fmt, variable=var).pack(anchor="w", pady=1)
        chosen = []
        btns = ttk.Frame(frm)
        btns.pack(anchor="e", pady=(8, 0))
        ttk.Button(btns, text="Run", command=lambda: (chosen.append(var.get()), win.destroy())).pack(
            side="left", padx=4)
        ttk.Button(btns, text="Cancel", command=win.destroy).pack(side="left")
        win.grab_set()
        win.wait_window()
        return chosen[0] if chosen else None

    def _ai_run(self, jobs):
        if self._ai_busy:
            messagebox.showinfo("Busy", "An AI run is already in progress.")
            return
        instruction = self.ai_prompt.get("1.0", tk.END).strip()
        if not instruction:
            messagebox.showinfo("No instruction", "Enter or pick a review instruction first.")
            return

        # Always pick the storage file first; its extension is the export format.
        path = filedialog.asksaveasfilename(
            title="Store this run's results as… (file type = format)",
            defaultextension=".md",
            filetypes=[("Markdown", "*.md"), ("Text", "*.txt"), ("CSV", "*.csv"),
                       ("Excel", "*.xlsx"), ("JSON", "*.json")],
            initialfile=self._default_export_name("ai_review", "md"))
        if not path:
            self.status_var.set("AI run cancelled — no storage file chosen.")
            return
        self._ai_autosave_path = path
        self._ai_run_start = len(self._ai_results)
        self._ai_save_error_shown = False

        # Dynamically map the modern engine panel to active session choices
        provider_code = self.llm_choice_an.get().upper()
        service = "o" if provider_code == "O" else "b"
        service_label = "DLR Ollama" if provider_code == "O" else "BlaBla"
        model = get_selected_model(service_label)
        
        fmt = self.ai_format.get() or "Plain text"
        directive = _AI_OUTPUT_FORMATS.get(fmt, "")

        self._ai_busy = True
        self._ai_set_busy_buttons("disabled")
        self._ai_q = queue.Queue()
        self._queue_run_snapshot()
        self._ai_progress_start(len(jobs))
        self._ai_log(f"=== Free-form review: {len(jobs)} section(s) via {service_label} "
                     f"({model or 'default'}), answer as {fmt}, "
                     f"store: {os.path.basename(path)} ===")
        self._ai_append(f"\n=== Running on {len(jobs)} section(s) via {service_label}"
                        f" ({model or 'default'}) — answer as {fmt} ===\n")
        threading.Thread(target=self._ai_worker,
                         args=(jobs, instruction, service, service_label, model,
                               fmt, directive),
                         daemon=True).start()
        self.root.after(150, self._ai_poll)

    def _ai_worker(self, jobs, instruction, service, service_label, model,
                   fmt, directive):
        try:
            from data_extraction.ai_utils.llm_utils import llm_call
        except Exception as exc:  # noqa: BLE001
            self._ai_q.put(("error", f"Could not load LLM utilities: {exc}"))
            self._ai_q.put(None)
            return
        full_instruction = f"{instruction}\n{directive}" if directive else instruction
        for i, (title, text) in enumerate(jobs, 1):
            prompt = f"{full_instruction}\n\n---\nSECTION: {title}\n\n{text}"
            self._ai_q.put(("start", i))
            self._ai_log_bg(f"▶ [{i}/{len(jobs)}] Analyzing '{title}' "
                            f"({len(text or '')} chars): {self._ai_preview(text)}")
            self._ai_log_bg(f"→ Sent to LLM ({len(prompt)} chars): "
                            f"{self._ai_preview(prompt)}")
            try:
                self._ai_call_count += 1
                resp = llm_call(prompt, None, service, model)
            except Exception as exc:  # noqa: BLE001
                resp = f"[ERROR] {exc}"
            self._ai_log_bg(f"← Response ({len(str(resp or ''))} chars): "
                            f"{self._ai_preview(resp)}")
            self._ai_q.put(("result", {
                "title": title, "instruction": instruction, "response": resp,
                "service": service_label, "model": model or "", "format": fmt,
            }))
        self._ai_q.put(None)

    def _ai_poll(self):
        try:
            while True:
                item = self._ai_q.get_nowait()
                if item is None:
                    self._ai_busy = False
                    self._ai_set_busy_buttons("normal")
                    self._ai_progress_done()
                    done = f"AI review complete — {len(self._ai_results)} result(s) total"
                    saved = len(self._ai_results) - self._ai_run_start
                    if self._ai_autosave_path and saved:
                        done += f", {saved} saved to {os.path.basename(self._ai_autosave_path)}"
                        self._ai_note_output(self._ai_autosave_path)
                    self.status_var.set(done + ".")
                    self._ai_log(done + ".")
                    return
                kind, payload = item
                if kind == "error":
                    self._ai_append(f"[ERROR] {payload}\n")
                    self._ai_log(f"[ERROR] {payload}")
                elif kind == "start":
                    self._run_current = payload
                    self._queue_set_status(payload - 1, "running")
                    self._ai_progress_info_update()
                else:
                    self._ai_results.append(payload)
                    self._ai_append(f"\n## {payload['title']}\n{payload['response']}\n")
                    self._ai_autosave()  # keep the file current after every result
                    ok = not str(payload.get("response", "")).startswith("[ERROR]")
                    self._queue_set_status(self._run_current - 1,
                                           "done" if ok else "error")
                    self._ai_progress_step()
        except queue.Empty:
            pass
        self._ai_progress_info_update()   # live call counter while working
        self.root.after(150, self._ai_poll)

    def _ai_export(self):
        if not self._ai_results:
            messagebox.showinfo("No results", "Run an AI review first, then export.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".md",
            filetypes=[("Markdown", "*.md"), ("Text", "*.txt"), ("CSV", "*.csv"),
                       ("Excel", "*.xlsx"), ("JSON", "*.json")],
            initialfile=self._default_export_name("ai_review", "md"))
        if not path:
            return
        try:
            self._write_ai_results(path, self._ai_results)
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("Export failed", str(exc))
            return
        self.status_var.set(f"Exported {len(self._ai_results)} AI result(s) → {path}")
        messagebox.showinfo("Export complete", f"Wrote {len(self._ai_results)} result(s) to:\n{path}")

    @staticmethod
    def _write_ai_results(path, results):
        """Write free-form results; the extension picks the format. Raises on failure."""
        ext = os.path.splitext(path)[1].lower()
        columns = ["title", "service", "model", "format", "instruction", "response"]
        if ext == ".json":
            with open(path, "w", encoding="utf-8") as f:
                json.dump(results, f, indent=2, ensure_ascii=False)
        elif ext == ".csv":
            import csv
            with open(path, "w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=columns)
                w.writeheader()
                w.writerows(results)
        elif ext == ".xlsx":
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font
            wb = Workbook()
            ws = wb.active
            ws.title = "AI Review"
            ws.append([c.capitalize() for c in columns])
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for r in results:
                ws.append([str(r.get(c, "")) for c in columns])
            widths = {"A": 40, "B": 12, "C": 18, "D": 12, "E": 40, "F": 90}
            for col, width in widths.items():
                ws.column_dimensions[col].width = width
            wrap = Alignment(wrap_text=True, vertical="top")
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = wrap
            wb.save(path)
        else:  # markdown / text
            blocks = []
            for r in results:
                blocks.append(f"## {r['title']}\n"
                              f"*{r['service']} {r['model']}* — {r['instruction']}\n\n"
                              f"{r['response']}\n")
            with open(path, "w", encoding="utf-8") as f:
                f.write("\n".join(blocks))

    def _ai_autosave(self):
        """Rewrite the storage file with the current run's results so far —
        called after every LLM reply, so the file grows result by result."""
        path = self._ai_autosave_path
        results = self._ai_results[self._ai_run_start:]
        if not path or not results:
            return
        try:
            self._write_ai_results(path, results)
        except Exception as exc:  # noqa: BLE001
            if not self._ai_save_error_shown:
                self._ai_save_error_shown = True
                messagebox.showerror("Auto-save failed", f"{path}\n\n{exc}")
            self.status_var.set(f"Auto-save failed: {exc}")
        
    def _choose_model_action(self, provider_code):
        """
        Fetch the live list of available models for the selected provider
        ('O' = DLR Ollama, 'B' = Blablador), let the user pick one, and store
        it as the session model used by all subsequent LLM calls.
        """
        service = "DLR Ollama" if str(provider_code).upper() == "O" else "BlaBla"

        print(f"Fetching available {service} models...")
        try:
            models = list_available_models(service)
        except Exception as e:
            messagebox.showerror("Model list failed", f"Could not fetch models for {service}:\n{e}")
            return

        if not models:
            messagebox.showwarning("No models", f"No models returned for {service}.\nKeeping current: {get_selected_model(service)}")
            return

        current = get_selected_model(service)

        dialog = tk.Toplevel(self.root)
        dialog.title(f"Select {service} Model")
        dialog.geometry("560x400")
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text=f"Available {service} models (current: {current}):",
                  font=("Arial", 10, "bold")).pack(padx=10, pady=8, anchor="w")

        list_frame = ttk.Frame(dialog)
        list_frame.pack(fill="both", expand=True, padx=10, pady=5)
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical")
        listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set)
        scrollbar.config(command=listbox.yview)
        scrollbar.pack(side="right", fill="y")
        listbox.pack(side="left", fill="both", expand=True)

        for m in models:
            listbox.insert(tk.END, m)
        if current in models:
            idx = models.index(current)
            listbox.selection_set(idx)
            listbox.see(idx)

        def _on_confirm():
            sel = listbox.curselection()
            if not sel:
                messagebox.showinfo("No selection", "Please select a model, or close the dialog to keep the current one.")
                return
            set_selected_model(service, models[sel[0]])
            self._ui_state_save()
            dialog.destroy()

        ttk.Button(dialog, text="Use Selected Model", command=_on_confirm).pack(pady=10)

    def _ai_set_busy_buttons(self, state):
        self.ai_run_btn.configure(state=state)
        self.ai_batch_btn.configure(state=state)
        self.col_run_btn.configure(state=state)
        self.eval_run_btn.configure(state=state)
        if state == "disabled":
            self.col_rerun_btn.configure(state="disabled")
        else:
            self._col_rerun_btn_sync()

    def _col_failed_rows(self):
        """Rows of the last analysis whose cells hold [ERROR]/[unparsed]."""
        return [r for r in self._col_results
                if any(isinstance(v, str)
                       and (v.startswith("[ERROR]") or v.startswith("[unparsed]"))
                       for v in r.values())]

    def _col_rerun_btn_sync(self):
        """'Re-run failed' is only clickable when a finished run left failed
        rows behind (and its settings snapshot exists to repeat them with)."""
        ready = (not self._ai_busy and bool(self._col_run_ctx)
                 and bool(self._col_failed_rows()))
        self.col_rerun_btn.configure(state="normal" if ready else "disabled")

    # ---------------------------------------------- session-state persist -- #
    def _ui_state_save(self):
        """Persist the page's session choices (storage file, LLM service and
        models, auto-eval picker, Evaluation-tab ticks, window geometry and
        console sash) so the next launch starts where this one left off.
        Best-effort: a failure never disturbs the app."""
        try:
            sash = None
            try:
                if getattr(self._ai_vsplit, "_sash_placed", False):
                    sash = int(self._ai_vsplit.sashpos(0))
            except Exception:  # noqa: BLE001
                pass
            state = {
                "store_path": self._col_store_path,
                "llm_choice": self.llm_choice_an.get(),
                "models": {lbl: get_selected_model(lbl)
                           for lbl in ("BlaBla", "DLR Ollama")},
                "eval_auto": bool(self.eval_auto_var.get()),
                "col_eval_choices": {k: bool(v) for k, v
                                     in self._col_eval_choices.items()},
                "eval_metrics": {label: bool(var.get()) for label, var
                                 in self.eval_metric_vars.items()},
                "eval_uniq": bool(self.eval_uniq_var.get()),
                "eval_out_choice": self.eval_out_choice.get(),
                "eval_wb": self.eval_wb.get().strip(),
                "geometry": self.root.winfo_toplevel().geometry(),
                "sash": sash,
            }
            path = _ui_state_path()
            path.parent.mkdir(parents=True, exist_ok=True)
            with open(path, "w", encoding="utf-8") as f:
                json.dump(state, f, indent=2, ensure_ascii=False)
        except Exception:  # noqa: BLE001 - saving is best-effort
            pass

    def _ui_state_load(self):
        """Restore the choices saved by :meth:`_ui_state_save`. Every field
        is applied independently and guarded, so a stale or partial state
        file can never block the startup."""
        try:
            path = _ui_state_path()
            if not path.exists():
                return
            with open(path, encoding="utf-8") as f:
                state = json.load(f)
        except Exception:  # noqa: BLE001
            return
        if isinstance(state.get("store_path"), str) and state["store_path"]:
            self._col_store_path = state["store_path"]
        if state.get("llm_choice") in ("O", "B"):
            self.llm_choice_an.set(state["llm_choice"])
        models = state.get("models") or {}
        for lbl in ("BlaBla", "DLR Ollama"):
            if isinstance(models.get(lbl), str) and models[lbl]:
                try:
                    set_selected_model(lbl, models[lbl])
                except Exception:  # noqa: BLE001
                    pass
        if isinstance(state.get("eval_auto"), bool):
            self.eval_auto_var.set(state["eval_auto"])
        if isinstance(state.get("col_eval_choices"), dict):
            self._col_eval_choices.update(
                {str(k): bool(v) for k, v in state["col_eval_choices"].items()})
        if isinstance(state.get("eval_metrics"), dict):
            for label, val in state["eval_metrics"].items():
                if label in self.eval_metric_vars:
                    self.eval_metric_vars[label].set(bool(val))
            try:  # re-apply the backend gating (e.g. BERTScore availability)
                self._eval_backend_changed()
            except Exception:  # noqa: BLE001
                pass
        if isinstance(state.get("eval_uniq"), bool):
            self.eval_uniq_var.set(state["eval_uniq"])
        if state.get("eval_out_choice") in ("same", "new"):
            self.eval_out_choice.set(state["eval_out_choice"])
        wb = state.get("eval_wb")
        if isinstance(wb, str) and wb and os.path.exists(wb):
            self.eval_wb.delete(0, tk.END)
            self.eval_wb.insert(0, wb)
            try:
                self._eval_refresh_sheets()
            except Exception:  # noqa: BLE001
                pass
        geo = state.get("geometry")
        if isinstance(geo, str) and re.match(r"^\d+x\d+", geo):
            try:
                self.root.winfo_toplevel().geometry(geo)
            except Exception:  # noqa: BLE001
                pass
        if isinstance(state.get("sash"), int):
            self._ui_saved_sash = state["sash"]

    # ------------------------------------------------- AI column analysis -- #
    def _col_load_columns(self):
        """Restore the column definitions saved by a previous session, if any.
        A missing/broken config file just keeps the built-in defaults."""
        try:
            path = _col_config_path()
            if not path.exists():
                return
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
            if "columns" not in data:
                return
            cols = [(str(n), str(w)) for n, w in data["columns"] if str(n)]
            self._ai_columns = cols   # may be empty: "Clear all" also persists
            names = {n for n, _ in cols}
            self._col_uniq_checked = {
                str(n) for n in data.get("uniq_checked", []) if str(n) in names}
            # older configs have no analyze set: analyze everything then
            self._col_analyze_checked = ({
                str(n) for n in data["analyze_checked"] if str(n) in names}
                if "analyze_checked" in data else set(names))
        except Exception:  # noqa: BLE001 - never block startup on the config
            pass

    def _col_save_columns(self):
        """Persist the current column definitions (and their ✓ unique-element
        checks) so they are back after closing and reopening the studio."""
        try:
            path = _col_config_path()
            path.parent.mkdir(parents=True, exist_ok=True)
            with open(path, "w", encoding="utf-8") as f:
                json.dump({"columns": [list(c) for c in self._ai_columns],
                           "uniq_checked": sorted(self._col_uniq_checked),
                           "analyze_checked": sorted(self._col_analyze_checked)},
                          f, indent=2, ensure_ascii=False)
        except Exception as exc:  # noqa: BLE001 - saving is best-effort
            self.status_var.set(f"Could not save the column definitions: {exc}")

    def _col_analyze_cols(self):
        """The Run ✓-checked column definitions — what the next run analyzes."""
        return [(n, w) for n, w in self._ai_columns
                if n in self._col_analyze_checked]

    def _col_refresh_defs(self):
        """Redraw the column-definition list and the live prompt preview."""
        self.col_defs.delete(*self.col_defs.get_children())
        for name, what in self._ai_columns:
            an = "☑" if name in self._col_analyze_checked else "☐"
            uq = "☑" if name in self._col_uniq_checked else "☐"
            self.col_defs.insert("", "end", text=name, values=(an, uq, what))
        self._col_refresh_preview()

    def _col_prompt_text(self):
        lines = ["Analyze the regulatory section below and extract the following data.",
                 "Return ONLY one valid JSON object with exactly these keys "
                 "(no code fences, no commentary):"]
        for name, what in self._col_analyze_cols():
            lines.append(f'- "{name}": {what or "extract this value"}')
        lines.append('Use "" for any value the section does not contain.')
        lines.append('Use ";" for multiple values the section does contain.')
        return "\n".join(lines)

    def _col_prompt_head(self):
        """Prompt head actually sent: the preview text above the "---" marker,
        so user edits made in the preview box are honoured. Falls back to the
        generated prompt when the box was emptied or still shows the
        no-columns placeholder."""
        text = self.col_preview.get("1.0", "end-1c")
        head = text.split("\n---\n", 1)[0].strip()
        if not head or head.startswith("Add at least one column"):
            return self._col_prompt_text()
        return head

    def _col_on_select(self, event=None):
        sel = self.col_defs.selection()
        if not sel:
            return
        name = self.col_defs.item(sel[0], "text")
        values = self.col_defs.item(sel[0], "values") or ("", "", "")
        what = values[2] if len(values) > 2 else ""
        self.col_name.delete(0, tk.END)
        self.col_name.insert(0, name)
        self.col_what.delete(0, tk.END)
        self.col_what.insert(0, what)

    def _col_defs_click(self, event):
        """Toggle a column's Run ✓ (analyze) or Uniq ✓ cell on click."""
        if self.col_defs.identify("region", event.x, event.y) != "cell":
            return None
        col_id = self.col_defs.identify_column(event.x)
        if col_id not in ("#1", "#2"):
            return None
        iid = self.col_defs.identify_row(event.y)
        if not iid:
            return None
        name = self.col_defs.item(iid, "text")
        checked = (self._col_analyze_checked if col_id == "#1"
                   else self._col_uniq_checked)
        if name in checked:
            checked.discard(name)
        else:
            checked.add(name)
        self.col_defs.set(iid, "an" if col_id == "#1" else "uniq",
                          "☑" if name in checked else "☐")
        self._col_save_columns()
        if col_id == "#1":   # the prompt only contains analyzed columns
            self._col_refresh_preview()
        return "break"

    def _col_refresh_preview(self):
        """Rebuild only the prompt preview (keeps the defs list untouched)."""
        self.col_preview.delete("1.0", tk.END)
        if self._col_analyze_cols():
            self.col_preview.insert("1.0", self._col_prompt_text() + _PREVIEW_TAIL)
        elif self._ai_columns:
            self.col_preview.insert(
                "1.0", "No column has its Run ✓ ticked — tick at least one "
                       "to build the prompt.")
        else:
            self.col_preview.insert("1.0", "Add at least one column to build the prompt.")

    def _col_toggle_analyze_all(self):
        """Run ✓ heading click: check every column, or clear if all checked."""
        names = {name for name, _ in self._ai_columns}
        if names and names <= self._col_analyze_checked:
            self._col_analyze_checked -= names
        else:
            self._col_analyze_checked |= names
        self._col_save_columns()
        self._col_refresh_defs()

    def _col_toggle_uniq_all(self):
        """Uniq ✓ heading click: check every column, or clear if all checked."""
        names = {name for name, _ in self._ai_columns}
        if names and names <= self._col_uniq_checked:
            self._col_uniq_checked -= names
        else:
            self._col_uniq_checked |= names
        self._col_save_columns()
        self._col_refresh_defs()

    def _col_add_update(self):
        name = self.col_name.get().strip()
        what = self.col_what.get().strip()
        if not name:
            messagebox.showinfo("Column name missing", "Give the column a name first.")
            return
        for i, (existing, _) in enumerate(self._ai_columns):
            if existing == name:
                self._ai_columns[i] = (name, what)
                break
        else:
            self._ai_columns.append((name, what))
            self._col_uniq_checked.add(name)   # new columns collect uniques by default
            self._col_analyze_checked.add(name)  # …and are analyzed by default
        self._col_save_columns()
        self._col_refresh_defs()

    def _col_add_entity_column(self):
        """Add (or refresh) the predefined 'Specific entities' column: the
        aviation Reference-System-Process-Personal-Quantity chain extraction
        (see ai_utils.entity_chains). Its chains are auto-parsed into an
        'Entities <run sheet>' component sheet during every run."""
        from .entity_chains import ENTITY_COLUMN, ENTITY_PROMPT
        for i, (existing, _) in enumerate(self._ai_columns):
            if existing == ENTITY_COLUMN:
                self._ai_columns[i] = (ENTITY_COLUMN, ENTITY_PROMPT)
                break
        else:
            self._ai_columns.append((ENTITY_COLUMN, ENTITY_PROMPT))
        self._col_analyze_checked.add(ENTITY_COLUMN)
        # chains are parsed into their own Entities sheet; whole-chain
        # unique collection is rarely useful, so it starts unticked
        self._col_uniq_checked.discard(ENTITY_COLUMN)
        self._col_save_columns()
        self._col_refresh_defs()
        self.status_var.set(f"'{ENTITY_COLUMN}' column added — its chains are "
                            "parsed into an 'Entities <run sheet>' sheet "
                            "automatically.")

    def _col_remove(self):
        names = {self.col_defs.item(iid, "text") for iid in self.col_defs.selection()}
        if not names:
            return
        self._ai_columns = [(n, w) for n, w in self._ai_columns if n not in names]
        self._col_uniq_checked -= names
        self._col_analyze_checked -= names
        self._col_save_columns()
        self._col_refresh_defs()

    def _col_clear(self):
        self._ai_columns = []
        self._col_uniq_checked.clear()
        self._col_analyze_checked.clear()
        self._col_save_columns()
        self._col_refresh_defs()

    def _col_run(self):
        if self._ai_busy:
            messagebox.showinfo("Busy", "An AI run is already in progress.")
            return
        if not self._ai_columns:
            messagebox.showinfo("No columns", "Define at least one column to extract.")
            return
        if not self._ai_batch:
            messagebox.showinfo("Empty queue",
                                "Add sections to the queue first (current node or checked nodes).")
            return
        jobs = list(self._ai_batch)
        # Only the Run ✓-checked columns are analyzed in this batch.
        col_defs = self._col_analyze_cols()
        if not col_defs:
            messagebox.showinfo(
                "No columns selected",
                "Every column's Run ✓ is unticked — tick at least one "
                "column to analyze.")
            return
        columns = [name for name, _ in col_defs]
        # The preview box is editable — send what it shows, not the template.
        prompt_head = self._col_prompt_head()

        # One combined call per section, or one call per column value?
        per_col = self._col_ask_call_mode()
        if per_col is None:
            self.status_var.set("Column analysis cancelled.")
            return

        # Always pick the storage file first. Re-choosing the same .xlsx appends
        # this run as a new snapshot sheet instead of overwriting.
        opts = {}
        if self._col_store_path:
            opts["initialdir"] = os.path.dirname(self._col_store_path)
            opts["initialfile"] = os.path.basename(self._col_store_path)
        else:
            opts["initialfile"] = self._default_export_name("ai_columns", "xlsx")
        store = filedialog.asksaveasfilename(
            title="Store analysis results in… (existing .xlsx gets this run as a new sheet)",
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx"), ("CSV", "*.csv"), ("JSON", "*.json")],
            confirmoverwrite=False, **opts)
        if not store:
            self.status_var.set("Column analysis cancelled — no storage file chosen.")
            return
        self._col_store_path = store
        # With the auto-evaluate box ticked, ask which evaluations should run
        # on each section result as soon as it is saved ("Skip evaluation"
        # opts out for this batch).
        auto_eval = False
        if self.eval_auto_var.get() and store.lower().endswith(".xlsx"):
            auto_eval = self._col_ask_auto_eval()
        self._col_run_auto_eval = auto_eval
        # Snapshot the picker's choices so mid-run changes don't affect this batch.
        self._col_run_eval_metrics = [
            n for label, names in _EVAL_METRIC_OPTIONS
            if self._col_eval_choices.get(label, True) for n in names
        ] if auto_eval else []
        self._col_run_eval_uniq = auto_eval and self._col_eval_choices.get("__uniq__", True)
        self._col_eval_entries = []     # per-section evaluations of this run
        self._col_rows_saved = 0        # file/sheet is created with the 1st row
        self._col_run_sheet = None
        self._col_save_error_shown = False
        # Only ✓-checked columns get unique-element collection this run.
        self._col_run_uniq_cols = [c for c in columns if c in self._col_uniq_checked]
        # Reference texts of this run's sections, for the evaluation afterwards.
        self._col_run_refs = dict(jobs)

        # Dynamically map modern engine panel values to active session configurations
        provider_code = self.llm_choice_an.get().upper()
        service = "o" if provider_code == "O" else "b"
        service_label = "DLR Ollama" if provider_code == "O" else "BlaBla"
        model = get_selected_model(service_label)

        # New run: rebuild the table for the current column set.
        self._col_results = []
        self._col_table_cols = ["Section"] + columns
        self.col_table.delete(*self.col_table.get_children())
        self.col_table.configure(columns=self._col_table_cols)
        for c in self._col_table_cols:
            self.col_table.heading(c, text=c)
            self.col_table.column(c, width=140 if c == "Section" else 180, anchor="w")

        # Settings snapshot so "Re-run failed" can repeat exactly this run.
        self._col_run_ctx = {"col_defs": col_defs, "prompt_head": prompt_head,
                             "per_col": per_col, "service": service,
                             "model": model}
        self._col_rerun = False
        self._ui_state_save()

        self._ai_busy = True
        self._ai_set_busy_buttons("disabled")
        self._col_pause_ev.clear()
        self._col_stop_ev.clear()
        self.col_pause_btn.configure(state="normal", text="Pause")
        self.col_stop_btn.configure(state="normal")
        self._col_q = queue.Queue()
        self._queue_run_snapshot()
        self._ai_progress_start(len(jobs))
        self._ai_log(f"=== Column analysis: {len(jobs)} section(s) → "
                     f"columns [{', '.join(columns)}] via {service_label} "
                     f"({model or 'default'}), "
                     + ("one LLM call per column value, "
                        if per_col else "one LLM call per section, ")
                     + f"store: {os.path.basename(store)} ===")
        self._ai_log("Auto-evaluation: "
                     + (f"[{', '.join(self._col_run_eval_metrics)}]"
                        + (" + unique values" if self._col_run_eval_uniq else "")
                        if auto_eval else "off"))
        self.status_var.set(f"Analyzing {len(jobs)} section(s) into {len(columns)} column(s)…")
        threading.Thread(target=self._col_worker,
                         args=(jobs, col_defs, prompt_head, service, model, per_col),
                         daemon=True).start()
        self.root.after(150, self._col_poll)

    def _col_ask_call_mode(self):
        """Modal: one combined LLM call per section (all columns in one JSON)
        or one LLM call per column value. Returns True for per-column,
        False for per-section, None if the user cancelled."""
        win = tk.Toplevel(self.root)
        win.title("LLM calls")
        win.transient(self.root.winfo_toplevel())
        win.resizable(False, False)
        frm = ttk.Frame(win, padding=12)
        frm.pack(fill="both", expand=True)
        ttk.Label(frm, text="How should the values be requested from the LLM?"
                  ).pack(anchor="w")
        mode = tk.StringVar(value="section")
        ttk.Radiobutton(
            frm, variable=mode, value="section",
            text="One call per section — all columns in a single JSON answer "
                 "(faster, fewer calls)").pack(anchor="w", pady=(8, 2))
        ttk.Radiobutton(
            frm, variable=mode, value="column",
            text="One call per column value — each column of a section is a "
                 "separate LLM call (slower, more focused)").pack(anchor="w")
        result = {"value": None}
        btns = ttk.Frame(frm)
        btns.pack(fill="x", pady=(12, 0))

        def _ok():
            result["value"] = mode.get() == "column"
            win.destroy()
        ttk.Button(btns, text="OK — start the analysis", command=_ok
                   ).pack(side="left")
        ttk.Button(btns, text="Cancel", command=win.destroy
                   ).pack(side="left", padx=8)
        win.grab_set()
        win.wait_window()
        return result["value"]

    def _col_rerun_failed(self):
        """Re-analyze only the failed rows ([ERROR]/[unparsed]) of the last
        run with its snapshotted settings, updating their existing rows in
        the table and the run sheet instead of appending new ones."""
        if self._ai_busy:
            return
        ctx = self._col_run_ctx
        failed = self._col_failed_rows()
        if not ctx or not failed:
            self._col_rerun_btn_sync()
            return
        jobs, missing = [], []
        for r in failed:
            title = str(r.get("Section", ""))
            text = self._col_run_refs.get(title)
            if text is None:
                missing.append(title)
            else:
                jobs.append((title, text))
        if missing:
            self._ai_log("[WARN] no section text for: " + ", ".join(missing))
        if not jobs:
            messagebox.showinfo("Re-run failed",
                                "The failed sections' texts are not available "
                                "any more — run the analysis again instead.")
            return
        self._col_rerun = True
        self._ai_busy = True
        self._ai_set_busy_buttons("disabled")
        self._col_pause_ev.clear()
        self._col_stop_ev.clear()
        self.col_pause_btn.configure(state="normal", text="Pause")
        self.col_stop_btn.configure(state="normal")
        self._col_q = queue.Queue()
        self._queue_snapshot_for([t for t, _ in jobs])
        self._ai_progress_start(len(jobs))
        self._ai_log(f"=== Re-running {len(jobs)} failed section(s): "
                     + ", ".join(f"'{t}'" for t, _ in jobs) + " ===")
        self.status_var.set(f"Re-running {len(jobs)} failed section(s)…")
        threading.Thread(target=self._col_worker,
                         args=(jobs, ctx["col_defs"], ctx["prompt_head"],
                               ctx["service"], ctx["model"], ctx["per_col"]),
                         daemon=True).start()
        self.root.after(150, self._col_poll)

    def _col_wait_if_paused(self):
        """Worker helper: block while paused; False = the run was stopped."""
        while self._col_pause_ev.is_set() and not self._col_stop_ev.is_set():
            time.sleep(0.2)
        return not self._col_stop_ev.is_set()

    @staticmethod
    def _col_single_prompt(name, what):
        """Prompt head asking for exactly one column value."""
        return (
            "Analyze the regulatory section below and extract one value.\n"
            "Return ONLY one valid JSON object with exactly this key "
            "(no code fences, no commentary):\n"
            f'- "{name}": {what or "extract this value"}\n'
            'Use "" if the section does not contain it.\n'
            'Use ";" for multiple values the section does contain.')

    def _col_worker(self, jobs, col_defs, prompt_head, service, model, per_col):
        try:
            from data_extraction.ai_utils.llm_utils import llm_call
        except Exception as exc:  # noqa: BLE001
            self._col_q.put(("error", f"Could not load LLM utilities: {exc}"))
            self._col_q.put(None)
            return
        columns = [name for name, _ in col_defs]

        def _norm(v):
            if isinstance(v, (list, tuple)):
                return "; ".join(str(x) for x in v)
            if isinstance(v, dict):
                return json.dumps(v, ensure_ascii=False)
            return "" if v is None else str(v)

        def _call_parsed(prompt, tag=""):
            """llm_call that re-sends the prompt up to _JSON_RETRIES extra
            times while the reply is not parsable JSON. Returns
            (parsed dict or None, last raw response); call errors propagate."""
            resp = None
            for attempt in range(_JSON_RETRIES + 1):
                if attempt:
                    if not self._col_wait_if_paused():
                        break   # stop requested mid-retry
                    self._ai_log_bg(f"↻{tag} response was not valid JSON — "
                                    f"retry {attempt}/{_JSON_RETRIES}")
                self._ai_call_count += 1
                resp = llm_call(prompt, None, service, model)
                self._ai_log_bg(f"←{tag} Response ({len(str(resp or ''))} chars): "
                                f"{self._ai_preview(resp)}")
                parsed = self._parse_llm_json(resp)
                if parsed is not None:
                    return parsed, resp
            return None, resp

        for i, (title, text) in enumerate(jobs, 1):
            if not self._col_wait_if_paused():
                self._col_q.put(("stopped", i - 1))
                break
            self._col_q.put(("start", i))
            self._ai_log_bg(f"▶ [{i}/{len(jobs)}] Analyzing '{title}' "
                            f"({len(text or '')} chars): {self._ai_preview(text)}")
            row = {"Section": title}
            if per_col:
                stopped = False
                for name, what in col_defs:
                    if not self._col_wait_if_paused():
                        stopped = True
                        break
                    if _is_entity_column(name):
                        # The Specific-entities prompt asks for a bare
                        # formatted string, not JSON — send it as-is and
                        # keep the raw reply.
                        prompt = f"{what}\n\n---\nSECTION: {title}\n\n{text}"
                        self._ai_log_bg(f"→ [column '{name}'] Sent to LLM "
                                        f"({len(prompt)} chars): "
                                        f"{self._ai_preview(prompt)}")
                        try:
                            self._ai_call_count += 1
                            resp = llm_call(prompt, None, service, model)
                        except Exception as exc:  # noqa: BLE001
                            row[name] = f"[ERROR] {exc}"
                            self._ai_log_bg(f"← [column '{name}'] LLM call "
                                            f"failed: {exc}")
                            continue
                        self._ai_log_bg(f"← [column '{name}'] Response "
                                        f"({len(str(resp or ''))} chars): "
                                        f"{self._ai_preview(resp)}")
                        row[name] = _norm(resp).strip()
                        continue
                    prompt = (f"{self._col_single_prompt(name, what)}"
                              f"\n\n---\nSECTION: {title}\n\n{text}")
                    self._ai_log_bg(f"→ [column '{name}'] Sent to LLM "
                                    f"({len(prompt)} chars): {self._ai_preview(prompt)}")
                    try:
                        parsed, resp = _call_parsed(prompt, f" [column '{name}']")
                    except Exception as exc:  # noqa: BLE001
                        row[name] = f"[ERROR] {exc}"
                        self._ai_log_bg(f"← [column '{name}'] LLM call failed: {exc}")
                        continue
                    if parsed is None:
                        row[name] = f"[unparsed] {_norm(resp).strip()}"
                        self._ai_log_bg(
                            f"   [column '{name}'] still not valid JSON after "
                            f"{_JSON_RETRIES} retries — marked [unparsed]")
                    else:
                        # accept the expected key, else the only/first value
                        v = parsed.get(name)
                        if v is None and parsed:
                            v = next(iter(parsed.values()))
                        row[name] = _norm(v)
                if stopped:
                    self._col_q.put(("stopped", i - 1))
                    break
            else:
                prompt = f"{prompt_head}\n\n---\nSECTION: {title}\n\n{text}"
                self._ai_log_bg(f"→ Sent to LLM ({len(prompt)} chars): "
                                f"{self._ai_preview(prompt)}")
                try:
                    parsed, resp = _call_parsed(prompt)
                except Exception as exc:  # noqa: BLE001
                    parsed = {columns[0]: f"[ERROR] {exc}"}
                    self._ai_log_bg(f"← LLM call failed: {exc}")
                else:
                    if parsed is None:
                        parsed = {columns[0]: f"[unparsed] {resp}"}
                        self._ai_log_bg(f"   still not valid JSON after "
                                        f"{_JSON_RETRIES} retries — kept as [unparsed]")
                for c in columns:
                    row[c] = _norm(parsed.get(c, ""))
            self._col_q.put(("row", row))
        self._col_q.put(None)

    @staticmethod
    def _parse_llm_json(resp):
        """Best-effort: pull the first {...} object out of an LLM reply."""
        if not resp or not isinstance(resp, str):
            return None
        start, end = resp.find("{"), resp.rfind("}")
        if start == -1 or end <= start:
            return None
        try:
            parsed = json.loads(resp[start:end + 1])
        except (ValueError, TypeError):
            return None
        return parsed if isinstance(parsed, dict) else None

    def _col_poll(self):
        try:
            while True:
                item = self._col_q.get_nowait()
                if item is None:
                    self._ai_busy = False
                    self.col_pause_btn.configure(state="disabled", text="Pause")
                    self.col_stop_btn.configure(state="disabled")
                    self._ai_progress_done()
                    what = ("Failed-row re-run" if self._col_rerun
                            else "Column analysis")
                    done = what + (" stopped" if self._col_stop_ev.is_set()
                                   else " complete")
                    done += f" — {len(self._col_results)} row(s)"
                    if self._col_store_path and self._col_rows_saved:
                        done += f", saved to {os.path.basename(self._col_store_path)}"
                        if self._col_run_sheet:
                            done += f" (sheet '{self._col_run_sheet}')"
                        self._ai_note_output(self._col_store_path)
                    self.status_var.set(done + ".")
                    self._ai_log(done + ".")
                    # Rows were evaluated as they were saved; finish up.
                    self._col_eval_finish()
                    self._col_rerun = False
                    self._ai_set_busy_buttons("normal")  # syncs Re-run failed
                    if self._col_failed_rows():
                        self._ai_log(f"⚠ {len(self._col_failed_rows())} row(s) "
                                     "still hold [ERROR]/[unparsed] — "
                                     "'Re-run failed' repeats only those.")
                    return
                kind, payload = item
                if kind == "error":
                    messagebox.showerror("Column analysis", payload)
                    self._ai_log(f"[ERROR] {payload}")
                elif kind == "stopped":
                    self._ai_log(f"■ Analysis stopped by the user after "
                                 f"{payload} completed section(s).")
                elif kind == "start":
                    self._run_current = payload
                    self._queue_set_status(payload - 1, "running")
                    self._ai_progress_info_update()
                else:
                    if self._col_rerun:
                        # Replace the failed row (results + table) in place.
                        title = payload.get("Section", "")
                        for i, r in enumerate(self._col_results):
                            if r.get("Section") == title:
                                self._col_results[i] = payload
                                break
                        else:
                            self._col_results.append(payload)
                        for item in self.col_table.get_children():
                            vals = self.col_table.item(item, "values")
                            if vals and vals[0] == title:
                                self.col_table.item(item, values=[
                                    payload.get(c, "")
                                    for c in self._col_table_cols])
                                break
                    else:
                        self._col_results.append(payload)
                        self.col_table.insert("", "end", values=[
                            payload.get(c, "") for c in self._col_table_cols])
                    self._col_save_row(payload)
                    ok = not any(isinstance(v, str)
                                 and (v.startswith("[ERROR]")
                                      or v.startswith("[unparsed]"))
                                 for v in payload.values())
                    self._queue_set_status(self._run_current - 1,
                                           "done" if ok else "error")
                    self._ai_progress_step()
        except queue.Empty:
            pass
        self._ai_progress_info_update()   # live call counter while working
        self.root.after(150, self._col_poll)

    def _col_pause_resume(self):
        """Toggle the pause of the running analysis (before the next LLM call)."""
        if not self._ai_busy:
            return
        if self._col_pause_ev.is_set():
            self._col_pause_ev.clear()
            self.col_pause_btn.configure(text="Pause")
            self.ai_paused_lbl.configure(text="")
            self._ai_log("▶ Analysis resumed.")
            self.status_var.set("Column analysis resumed.")
        else:
            self._col_pause_ev.set()
            self.col_pause_btn.configure(text="Resume")
            self.ai_paused_lbl.configure(text="⏸ PAUSED")
            self._ai_log("⏸ Analysis paused — the current LLM call still "
                         "finishes, then the run waits.")
            self.status_var.set("Column analysis paused.")

    def _col_stop(self):
        """Stop the running analysis after the current LLM call; already
        saved rows (and their evaluations) are kept."""
        if not self._ai_busy:
            return
        self._col_stop_ev.set()
        self._col_pause_ev.clear()   # unblock a paused worker so it can exit
        self.col_stop_btn.configure(state="disabled")
        self.col_pause_btn.configure(state="disabled", text="Pause")
        self.ai_paused_lbl.configure(text="")
        self._ai_log("■ Stop requested — finishing the current LLM call…")
        self.status_var.set("Stopping the column analysis…")

    def _col_show_row(self, event=None):
        sel = self.col_table.selection()
        if not sel or not self._col_table_cols:
            return
        values = self.col_table.item(sel[0], "values")
        win = tk.Toplevel(self.root)
        win.title("Row details")
        win.transient(self.root.winfo_toplevel())
        txt = scrolledtext.ScrolledText(win, wrap="word", width=90, height=24)
        txt.pack(fill="both", expand=True, padx=8, pady=8)
        for col, val in zip(self._col_table_cols, values):
            txt.insert(tk.END, f"■ {col}\n{val}\n\n")
        txt.configure(state="disabled")

    @staticmethod
    def _col_write_store(path, cols, rows):
        """Write one analysis snapshot. An existing .xlsx gets a new sheet per
        run (accumulating workbook); CSV/JSON hold the latest batch. Returns
        the sheet name used for .xlsx, else None. Raises on failure."""
        ext = os.path.splitext(path)[1].lower()
        if ext == ".json":
            with open(path, "w", encoding="utf-8") as f:
                json.dump(rows, f, indent=2, ensure_ascii=False)
            return None
        if ext == ".csv":
            import csv
            with open(path, "w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=cols)
                w.writeheader()
                w.writerows(rows)
            return None
        # Excel: one snapshot sheet per analysis run
        from datetime import datetime

        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
        if os.path.exists(path):
            wb = load_workbook(path)
            ws = wb.create_sheet()
        else:
            wb = Workbook()
            ws = wb.active
        stamp = datetime.now().strftime("%Y-%m-%d %H.%M.%S")
        run_no = sum(1 for s in wb.sheetnames if s.startswith("Run ")) + 1
        ws.title = f"Run {run_no} {stamp}"[:31]
        ws.append(cols)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for r in rows:
            ws.append([str(r.get(c, "")) for c in cols])
        ws.column_dimensions["A"].width = 40
        for i in range(2, len(cols) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 50
        wrap = Alignment(wrap_text=True, vertical="top")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap
        wb.save(path)
        return ws.title

    def _col_save_row(self, row):
        """Write one finished section to the storage file immediately: the file
        (and, for .xlsx, the run's snapshot sheet) is created with the first
        row, then updated and saved again for every further row."""
        path = self._col_store_path
        if not path:
            return
        cols = self._col_table_cols
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext == ".json":
                with open(path, "w", encoding="utf-8") as f:
                    json.dump(self._col_results, f, indent=2, ensure_ascii=False)
            elif ext == ".csv":
                import csv
                if self._col_rerun:
                    # replaced rows: rewrite the whole file from the results
                    with open(path, "w", newline="", encoding="utf-8") as f:
                        w = csv.DictWriter(f, fieldnames=cols)
                        w.writeheader()
                        w.writerows(self._col_results)
                else:
                    mode = "w" if self._col_rows_saved == 0 else "a"
                    with open(path, mode, newline="", encoding="utf-8") as f:
                        w = csv.DictWriter(f, fieldnames=cols)
                        if self._col_rows_saved == 0:
                            w.writeheader()
                        w.writerow(row)
            else:  # .xlsx
                from datetime import datetime

                from openpyxl import Workbook, load_workbook
                from openpyxl.styles import Alignment, Font
                from openpyxl.utils import get_column_letter
                if self._col_rerun and self._col_run_sheet:
                    # Re-run of a failed row: overwrite its existing sheet
                    # row (matched by the Section cell) instead of appending.
                    wb = load_workbook(path)
                    ws = wb[self._col_run_sheet]
                    wrap = Alignment(wrap_text=True, vertical="top")
                    title = str(row.get("Section", ""))
                    for r in range(2, ws.max_row + 1):
                        if str(ws.cell(row=r, column=1).value or "") == title:
                            for j, c in enumerate(cols, 1):
                                cell = ws.cell(row=r, column=j,
                                               value=str(row.get(c, "")))
                                cell.alignment = wrap
                            break
                    else:  # row vanished from the sheet: append it back
                        ws.append([str(row.get(c, "")) for c in cols])
                        for cell in ws[ws.max_row]:
                            cell.alignment = wrap
                    wb.save(path)
                else:
                    if self._col_rows_saved == 0:
                        if os.path.exists(path):
                            wb = load_workbook(path)
                            ws = wb.create_sheet()
                        else:
                            wb = Workbook()
                            ws = wb.active
                        stamp = datetime.now().strftime("%Y-%m-%d %H.%M.%S")
                        run_no = sum(1 for s in wb.sheetnames
                                     if s.startswith("Run ")) + 1
                        ws.title = f"Run {run_no} {stamp}"[:31]
                        self._col_run_sheet = ws.title
                        ws.append(cols)
                        for cell in ws[1]:
                            cell.font = Font(bold=True)
                        ws.column_dimensions["A"].width = 40
                        for i in range(2, len(cols) + 1):
                            ws.column_dimensions[get_column_letter(i)].width = 50
                    else:
                        wb = load_workbook(path)
                        ws = wb[self._col_run_sheet]
                    ws.append([str(row.get(c, "")) for c in cols])
                    wrap = Alignment(wrap_text=True, vertical="top")
                    for cell in ws[ws.max_row]:
                        cell.alignment = wrap
                    wb.save(path)
                self._col_update_uniques(path)
                self._col_update_entities(path)
                self._col_eval_saved_row(row)
            if not self._col_rerun:
                self._col_rows_saved += 1
            verb = "re-saved" if self._col_rerun else "saved"
            self.status_var.set(f"Row ({row.get('Section', '')}) {verb} → "
                                f"{os.path.basename(path)}")
            self._ai_log(f"💾 Row '{row.get('Section', '')}' {verb} → "
                         f"{os.path.basename(path)}"
                         + (f" (sheet '{self._col_run_sheet}')"
                            if self._col_run_sheet else ""))
        except Exception as exc:  # noqa: BLE001
            if not self._col_save_error_shown:
                self._col_save_error_shown = True
                messagebox.showerror(
                    "Row save failed",
                    f"{path}\n\n{exc}\n\nThe analysis continues; further save "
                    f"errors only go to the status bar.")
            self.status_var.set(f"Row save failed: {exc}")

    def _col_update_uniques(self, path):
        """Refresh the run's unique-elements sheet (element + count) via
        scripts/excel_file_utils after each saved row — but only for the
        columns whose ✓ checkbox was ticked when the run started."""
        data_cols = [c for c in self._col_run_uniq_cols if c != "Section"]
        if not data_cols or not self._col_run_sheet:
            return
        uniq_sheet = f"Uniq {self._col_run_sheet}"[:31]
        _excel_utils().save_unique_elements_to_new_sheet(
            path, data_cols, new_sheet_name=uniq_sheet,
            source_sheet=self._col_run_sheet)

    def _col_update_entities(self, path):
        """Auto extraction for the 'Specific entities' column: after each
        saved row, re-parse every chain of this run into an
        'Entities <run sheet>' sheet (one row per chain, one column per
        component). No-ops when the column is not part of the run; failures
        only reach the console/status bar."""
        if not self._col_run_sheet:
            return
        if not any(_is_entity_column(c) for c in self._col_table_cols):
            return
        try:
            from .entity_chains import extract_rows, write_entity_sheet
            rows = extract_rows(self._col_results)
            sheet = write_entity_sheet(
                path, rows, f"Entities {self._col_run_sheet}"[:31])
            self._ai_log(f"   → '{sheet}' refreshed ({len(rows)} chain(s))")
        except Exception as exc:  # noqa: BLE001 - never interrupt the analysis
            self.status_var.set(f"Entity-sheet update failed: {exc}")
            self._ai_log(f"[ERROR] entity-sheet update failed: {exc}")

    # ------------------------------------------------- Unique elements tab -- #
    def _build_ai_uniques_tab(self):
        """Standalone unique-element extractor: for any workbook + sheet, split
        the chosen columns on a chosen separator and write the unique values
        (element + count) into a new sheet of the *same* file. Independent of a
        Column-analysis run — it works on any existing Excel file."""
        uq = self._scrollable_tab("Unique elements")

        wb_wrap = ttk.LabelFrame(uq, text=" Excel file ", padding=4)
        wb_wrap.pack(fill="x")
        row = ttk.Frame(wb_wrap)
        row.pack(fill="x")
        ttk.Label(row, text="Workbook:").pack(side="left")
        self.uniq_wb = ttk.Entry(row)
        self.uniq_wb.pack(side="left", fill="x", expand=True, padx=4)
        ttk.Button(row, text="Browse…", command=self._uniq_browse_wb).pack(side="left")
        row2 = ttk.Frame(wb_wrap)
        row2.pack(fill="x", pady=(4, 0))
        ttk.Label(row2, text="Sheet:").pack(side="left")
        self.uniq_sheet = ttk.Combobox(row2, state="readonly", width=34, values=[])
        self.uniq_sheet.pack(side="left", padx=4)
        self.uniq_sheet.bind("<<ComboboxSelected>>", self._uniq_refresh_columns)
        ttk.Button(row2, text="Refresh sheets",
                   command=self._uniq_refresh_sheets).pack(side="left")

        col_wrap = ttk.LabelFrame(uq, text=" Columns to extract unique values from ",
                                  padding=4)
        col_wrap.pack(fill="x", pady=(6, 0))
        cbtns = ttk.Frame(col_wrap)
        cbtns.pack(fill="x")
        ttk.Button(cbtns, text="Select all",
                   command=lambda: self._uniq_set_all_cols(True)).pack(side="left")
        ttk.Button(cbtns, text="Clear",
                   command=lambda: self._uniq_set_all_cols(False)).pack(side="left", padx=4)
        # Checkboxes are (re)built here whenever a sheet is chosen.
        self.uniq_cols_frame = ttk.Frame(col_wrap)
        self.uniq_cols_frame.pack(fill="x", pady=(4, 0))
        self._uniq_cols_hint = ttk.Label(
            self.uniq_cols_frame, foreground="#666666",
            text="Pick a workbook and sheet, then Refresh sheets to list its columns.")
        self._uniq_cols_hint.pack(anchor="w")

        sep_wrap = ttk.LabelFrame(uq, text=" Item separator inside each cell ", padding=4)
        sep_wrap.pack(fill="x", pady=(6, 0))
        srow = ttk.Frame(sep_wrap)
        srow.pack(fill="x")
        ttk.Label(srow, text="Separator:").pack(side="left")
        self.uniq_sep = ttk.Combobox(srow, state="readonly", width=28,
                                     values=[lbl for lbl, _ in _UNIQ_SEPARATORS])
        self.uniq_sep.current(0)
        self.uniq_sep.pack(side="left", padx=4)
        self.uniq_sep.bind("<<ComboboxSelected>>", self._uniq_sep_changed)
        ttk.Label(srow, text="Custom:").pack(side="left", padx=(12, 0))
        self.uniq_sep_custom = ttk.Entry(srow, width=12)
        self.uniq_sep_custom.pack(side="left", padx=4)
        self.uniq_sep_custom.configure(state="disabled")
        orow = ttk.Frame(sep_wrap)
        orow.pack(fill="x", pady=(4, 0))
        self.uniq_case_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(orow, variable=self.uniq_case_var,
                        text="Case-insensitive (treat 'Foo' and 'foo' as one)").pack(
            side="left")
        ttk.Label(orow, text="Sort by:").pack(side="left", padx=(16, 0))
        self.uniq_sort = ttk.Combobox(orow, state="readonly", width=18,
                                      values=["Element (A→Z)", "Count (high→low)"])
        self.uniq_sort.current(0)
        self.uniq_sort.pack(side="left", padx=4)

        out_wrap = ttk.LabelFrame(uq, text=" New sheet (written into the same file) ",
                                  padding=4)
        out_wrap.pack(fill="x", pady=(6, 0))
        nrow = ttk.Frame(out_wrap)
        nrow.pack(fill="x")
        ttk.Label(nrow, text="Sheet name:").pack(side="left")
        self.uniq_out_name = ttk.Entry(nrow)
        self.uniq_out_name.pack(side="left", fill="x", expand=True, padx=4)
        ttk.Label(out_wrap, foreground="#666666", justify="left",
                  text="Columns are written as Unique_<col> / Count_<col> via "
                       "save_unique_elements_to_new_sheet (matching the run 'Uniq' "
                       "sheets).\nAn existing sheet of the same name is replaced. "
                       "Blank = 'Uniq <sheet>'.").pack(anchor="w", pady=(2, 0))

        # Optional extra pass: match the unique values of reference columns
        # against the section titles (column_evaluator.add_reference_sections —
        # the same function the run-sheet 'Uniq' pipeline uses).
        ref_wrap = ttk.LabelFrame(
            uq, text=" Reference columns → section titles (optional) ", padding=4)
        ref_wrap.pack(fill="x", pady=(6, 0))
        self.uniq_refmap_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            ref_wrap, variable=self.uniq_refmap_var,
            command=self._uniq_refmap_toggle,
            text="Also match reference columns (name contains 'reference') "
                 "against the section titles").pack(anchor="w")
        rrow = ttk.Frame(ref_wrap)
        rrow.pack(fill="x", pady=(4, 0))
        ttk.Label(rrow, text="Sections JSON:").pack(side="left")
        self.uniq_ref_json = ttk.Entry(rrow)
        self.uniq_ref_json.pack(side="left", fill="x", expand=True, padx=4)
        self.uniq_ref_browse = ttk.Button(rrow, text="Browse…",
                                          command=self._uniq_browse_ref_json)
        self.uniq_ref_browse.pack(side="left")
        ttk.Label(ref_wrap, foreground="#666666", justify="left",
                  text="Each unique value of those columns is matched against the "
                       "JSON's section titles (substring, either direction). The "
                       "new sheet keeps its Unique_/Count_ columns and gains "
                       "Section_<col> / Section_Count_<col> /\nSection_Matches_<col> "
                       "beside them (most-referenced first); every individual match "
                       "also goes to a separate 'Ref Map …' sheet.").pack(
            anchor="w", pady=(2, 0))
        self._uniq_refmap_toggle()

        # Standalone Specific-entities chain parser: any column holding
        # Reference-System-Process-Personal-Quantity chains is fanned out
        # into an 'Entities <sheet>' component sheet of the same file.
        ent_wrap = ttk.LabelFrame(
            uq, text=" Specific-entity chains → component sheet (standalone) ",
            padding=4)
        ent_wrap.pack(fill="x", pady=(6, 0))
        erow = ttk.Frame(ent_wrap)
        erow.pack(fill="x")
        ttk.Label(erow, text="Chain column:").pack(side="left")
        self.uniq_ent_col = ttk.Combobox(erow, state="readonly", width=30,
                                         values=[])
        self.uniq_ent_col.pack(side="left", padx=4)
        self.uniq_ent_btn = ttk.Button(erow, text="Parse chains → sheet",
                                       command=self._uniq_parse_entities)
        self.uniq_ent_btn.pack(side="left", padx=4)
        ttk.Label(ent_wrap, foreground="#666666", justify="left",
                  text="Parses Reference-System Info-Process-Personal-"
                       "QuantityValue chains (';'-separated, '#' = empty) "
                       "from the chosen column of the sheet above into an\n"
                       "'Entities <sheet>' sheet — one row per chain, one "
                       "column per component. Reference-less chains are "
                       "skipped; an existing sheet is replaced.").pack(
            anchor="w", pady=(2, 0))

        brow = ttk.Frame(uq)
        brow.pack(fill="x", pady=6)
        self.uniq_run_btn = ttk.Button(brow, text="Generate unique elements",
                                       command=self._uniq_generate)
        self.uniq_run_btn.pack(side="left")

    def _uniq_parse_entities(self):
        """Standalone chain parsing for the Unique-elements tab's workbook +
        sheet: parse the chosen column's Specific-entity chains into an
        'Entities <sheet>' component sheet of the same file."""
        path = self.uniq_wb.get().strip()
        sheet = self.uniq_sheet.get().strip()
        column = self.uniq_ent_col.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showinfo("No workbook", "Pick an Excel file first.")
            return
        if not sheet:
            messagebox.showinfo("No sheet", "Pick a sheet first (Refresh sheets).")
            return
        if not column:
            messagebox.showinfo(
                "No column", "Pick the column that holds the entity chains.")
            return
        try:
            from .entity_chains import entities_from_workbook
            summary = entities_from_workbook(path, sheet, column)
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("Entity chains", str(exc))
            self._ai_log(f"[ERROR] entity-chain parsing failed: {exc}")
            return
        self._ai_note_output(path)
        msg = (f"Parsed {summary['chains']} chain(s) from column '{column}' "
               f"({summary['rows']} row(s)) → sheet '{summary['sheet']}'")
        self.status_var.set(msg + ".")
        self._ai_log(f"=== Entity chains: {os.path.basename(path)} [{sheet}] "
                     f"— {msg} ===")
        messagebox.showinfo("Entity chains", msg + f"\n\nin {path}")

    def _uniq_refmap_toggle(self):
        """Enable the sections-JSON picker only while section matching is on."""
        state = "normal" if self.uniq_refmap_var.get() else "disabled"
        self.uniq_ref_json.configure(state=state)
        self.uniq_ref_browse.configure(state=state)

    def _uniq_browse_ref_json(self):
        path = filedialog.askopenfilename(
            title="Select the sections JSON holding the section titles",
            filetypes=[("JSON files", "*.json")])
        if path:
            self.uniq_ref_json.configure(state="normal")
            self.uniq_ref_json.delete(0, tk.END)
            self.uniq_ref_json.insert(0, path)

    def _uniq_browse_wb(self):
        path = filedialog.askopenfilename(
            title="Select an Excel workbook",
            filetypes=[("Excel workbook", "*.xlsx *.xlsm")])
        if path:
            self.uniq_wb.delete(0, tk.END)
            self.uniq_wb.insert(0, path)
            self._uniq_refresh_sheets()

    def _uniq_refresh_sheets(self):
        """List every sheet in the chosen workbook (not just 'Run …' sheets)."""
        path = self.uniq_wb.get().strip()
        sheets = []
        if path and os.path.exists(path):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(path, read_only=True)
                sheets = list(wb.sheetnames)
                wb.close()
            except Exception as exc:  # noqa: BLE001
                messagebox.showerror("Workbook", f"Could not read sheets:\n{exc}")
        self.uniq_sheet.configure(values=sheets)
        self.uniq_sheet.set(sheets[0] if sheets else "")
        self._uniq_refresh_columns()

    def _uniq_refresh_columns(self, event=None):
        """Rebuild the column checkboxes from the selected sheet's header row."""
        for child in self.uniq_cols_frame.winfo_children():
            child.destroy()
        self._uniq_col_vars = {}
        path = self.uniq_wb.get().strip()
        sheet = self.uniq_sheet.get().strip()
        if not (path and os.path.exists(path) and sheet):
            ttk.Label(self.uniq_cols_frame, foreground="#666666",
                      text="Pick a workbook and sheet to list its columns.").pack(
                anchor="w")
            self.uniq_ent_col.configure(values=[])
            self.uniq_ent_col.set("")
            return
        try:
            import pandas as pd
            head = pd.read_excel(path, sheet_name=sheet, nrows=0, engine="openpyxl")
            columns = [str(c) for c in head.columns]
        except Exception as exc:  # noqa: BLE001
            ttk.Label(self.uniq_cols_frame, foreground="#a00000",
                      text=f"Could not read columns: {exc}").pack(anchor="w")
            return
        # Feed the standalone chain parser too; preselect a column named
        # like the Specific-entities column when the sheet has one.
        self.uniq_ent_col.configure(values=columns)
        ent_default = next((c for c in columns if _is_entity_column(c)), "")
        self.uniq_ent_col.set(ent_default)
        if not columns:
            ttk.Label(self.uniq_cols_frame, foreground="#666666",
                      text="This sheet has no columns.").pack(anchor="w")
            return
        # Default the new-sheet name to 'Uniq <sheet>'. Follow the selected
        # sheet unless the user has typed their own name.
        current = self.uniq_out_name.get().strip()
        if not current or current == getattr(self, "_uniq_out_auto", None):
            self._uniq_out_auto = f"Uniq {sheet}"[:31]
            self.uniq_out_name.delete(0, tk.END)
            self.uniq_out_name.insert(0, self._uniq_out_auto)
        for i, col in enumerate(columns):
            var = tk.BooleanVar(value=(col != "Section"))
            self._uniq_col_vars[col] = var
            ttk.Checkbutton(self.uniq_cols_frame, text=col, variable=var).grid(
                row=i % 6, column=i // 6, sticky="w", padx=(0, 18))

    def _uniq_set_all_cols(self, value):
        for var in self._uniq_col_vars.values():
            var.set(value)

    def _uniq_sep_changed(self, event=None):
        """Enable the custom-separator field only when 'Custom…' is chosen."""
        _, sep = _UNIQ_SEPARATORS[self.uniq_sep.current()]
        self.uniq_sep_custom.configure(state="normal" if sep == "custom" else "disabled")

    def _uniq_selected_separator(self):
        """Resolve the separator choice to a value for
        save_unique_elements_to_new_sheet's ``separators`` argument: a list of
        candidates for auto-detect, or a single string for a fixed / custom
        separator. Returns ``False`` (a sentinel distinct from valid values)
        after warning when 'Custom…' is chosen but left empty."""
        idx = self.uniq_sep.current()
        _, sep = _UNIQ_SEPARATORS[idx if idx >= 0 else 0]
        if sep is None:
            return list(_UNIQ_AUTO_CANDIDATES)      # auto-detect among these
        if sep == "custom":
            text = self.uniq_sep_custom.get()
            if not text:
                messagebox.showinfo("Separator",
                                    "Type the custom separator, or pick another option.")
                return False
            return text
        return sep                                   # a single fixed separator

    def _uniq_generate(self):
        """Read the chosen sheet, split the chosen columns on the chosen
        separator and write Unique_/Count_ columns into a new sheet of the same
        file — delegating to scripts/excel_file_utils'
        save_unique_elements_to_new_sheet, in a worker thread so the UI stays
        responsive."""
        if self._uniq_busy or self._ai_busy or self._eval_busy:
            messagebox.showinfo("Busy", "Another run is already in progress.")
            return
        path = self.uniq_wb.get().strip()
        if not path or os.path.splitext(path)[1].lower() not in (".xlsx", ".xlsm") \
                or not os.path.exists(path):
            messagebox.showinfo("No workbook", "Pick an existing .xlsx/.xlsm file first.")
            return
        sheet = self.uniq_sheet.get().strip()
        if not sheet:
            messagebox.showinfo("No sheet", "Pick a sheet to read.")
            return
        columns = [c for c, v in self._uniq_col_vars.items() if v.get()]
        if not columns:
            messagebox.showinfo("No columns", "Tick at least one column.")
            return
        separators = self._uniq_selected_separator()
        if separators is False:      # empty custom separator (already warned)
            return
        case_insensitive = self.uniq_case_var.get()
        sort_by = "count" if self.uniq_sort.current() == 1 else "element"
        out_name = self._uniq_sanitize_sheet(
            self.uniq_out_name.get().strip() or f"Uniq {sheet}")
        if out_name == sheet:
            messagebox.showinfo("Sheet name",
                                "The new sheet name must differ from the source sheet.")
            return
        # Optional reference-column → section-title matching.
        ref_json = None
        if self.uniq_refmap_var.get():
            try:
                from data_extraction.evaluation.column_evaluator import (
                    is_reference_column)
            except Exception as exc:  # noqa: BLE001
                messagebox.showerror("Section matching",
                                     f"column_evaluator unavailable:\n{exc}")
                return
            ref_cols = [c for c in columns if is_reference_column(c)]
            if not ref_cols:
                messagebox.showinfo(
                    "No reference column",
                    "Section matching is ticked, but none of the ticked columns "
                    "has 'reference' in its name — tick a reference column, or "
                    "untick the section-matching box.")
                return
            ref_json = self.uniq_ref_json.get().strip()
            if not ref_json or not os.path.exists(ref_json):
                messagebox.showinfo(
                    "Sections JSON needed",
                    "Pick the sections JSON whose section titles the reference "
                    "values are matched against.")
                return

        self._uniq_busy = True
        self.uniq_run_btn.configure(state="disabled")
        self.ai_progress.configure(mode="indeterminate")
        self.ai_progress.start(80)
        sep_desc = ("auto-detect" if isinstance(separators, list)
                    else f"{separators!r}")
        self._ai_log(f"=== Unique elements: {os.path.basename(path)} [{sheet}], "
                     f"columns [{', '.join(columns)}], separator {sep_desc}"
                     f"{', case-insensitive' if case_insensitive else ''}, "
                     f"sort by {sort_by} → sheet '{out_name}' ===")
        if ref_json:
            self._ai_log(f"   reference columns [{', '.join(ref_cols)}] will be "
                         f"matched against the section titles in "
                         f"{os.path.basename(ref_json)}")
        self.status_var.set(f"Extracting unique values from {sheet}…")

        def work():
            try:
                summary = self._uniq_build_sheet(
                    path, sheet, columns, separators, case_insensitive,
                    sort_by, out_name, ref_json)
            except Exception as exc:  # noqa: BLE001
                self.root.after(0, lambda e=exc: self._uniq_done(None, e))
            else:
                self.root.after(0, lambda s=summary: self._uniq_done(s, None))

        threading.Thread(target=work, daemon=True).start()

    def _uniq_build_sheet(self, path, sheet, columns, separators,
                          case_insensitive, sort_by, out_name, ref_json=None):
        """Worker: delegate to save_unique_elements_to_new_sheet, then read the
        written sheet back for the per-column unique counts. When ``ref_json``
        is given, the reference columns' unique values are additionally matched
        against that JSON's section titles (column_evaluator.add_reference_sections
        — the same pass the run-sheet 'Uniq' pipeline runs). Runs off the main
        thread — no widget access here."""
        import pandas as pd

        ok = _excel_utils().save_unique_elements_to_new_sheet(
            path, columns, new_sheet_name=out_name, source_sheet=sheet,
            separators=separators, case_insensitive=case_insensitive,
            sort_by=sort_by)
        if not ok:
            raise RuntimeError(
                "save_unique_elements_to_new_sheet reported failure — none of "
                "the chosen columns were found, or the file could not be "
                "written (see the progress console for the printed reason).")
        # Count the unique values actually written, per column, for the summary.
        written = pd.read_excel(path, sheet_name=out_name, engine="openpyxl")
        per_col = {c[len("Unique_"):]: int(written[c].notna().sum())
                   for c in written.columns if c.startswith("Unique_")}
        summary = {"sheet": out_name, "rows": int(written.shape[0]),
                   "per_col": per_col, "path": path}
        if ref_json:
            from data_extraction.evaluation.column_evaluator import (
                add_reference_sections, ref_map_sheet_name, references_from_json)
            references = references_from_json(ref_json)
            if not references:
                raise RuntimeError(
                    f"No sections found in {os.path.basename(ref_json)} — the "
                    "reference values have no section titles to match against.")
            summary.update(add_reference_sections(
                path, out_name, references, columns,
                map_sheet=ref_map_sheet_name(out_name)))
            summary["ref_json"] = ref_json
            summary["ref_titles"] = len(references)
        return summary

    @staticmethod
    def _uniq_sanitize_sheet(name):
        """Excel sheet-name rules: <=31 chars, none of : \\ / ? * [ ]."""
        for ch in ':\\/?*[]':
            name = name.replace(ch, " ")
        return name.strip()[:31] or "Uniq"

    def _uniq_done(self, summary, exc):
        self._uniq_busy = False
        self.uniq_run_btn.configure(state="normal")
        self._ai_progress_done()
        if exc is not None:
            self.status_var.set(f"Unique-elements extraction failed: {exc}")
            self._ai_log(f"[ERROR] unique elements failed: {exc}")
            messagebox.showerror("Unique elements failed", str(exc))
            return
        self._ai_note_output(summary.get("path"))
        detail = ", ".join(f"{col}: {n}" for col, n in summary["per_col"].items())
        self.status_var.set(
            f"Unique elements written → sheet '{summary['sheet']}' "
            f"({detail}) in {os.path.basename(summary['path'])}")
        self._ai_log(f"=== Unique elements complete → '{summary['sheet']}' "
                     f"({detail}) ===")
        secs = summary.get("ref_sections") or {}
        if summary.get("ref_map_sheet"):
            self._ai_log(
                f"   → reference values matched against "
                f"{summary.get('ref_titles', 0)} section title(s): "
                + (", ".join(f"{c}: {len(t)} section(s) referenced"
                             for c, t in secs.items()) or "no section matched")
                + f" — Section_* columns added to '{summary['sheet']}', "
                  f"details → '{summary['ref_map_sheet']}'")
        # Refresh the sheet list so the new sheet shows up immediately.
        self._uniq_refresh_sheets()
        extra = ""
        if summary.get("ref_map_sheet"):
            extra = ("\n\nReference columns matched to section titles:\n"
                     + "\n".join(f"{c}: {len(t)} section(s) referenced"
                                 for c, t in secs.items())
                     + f"\n\nSection_* columns added to '{summary['sheet']}'; "
                       f"every match listed in '{summary['ref_map_sheet']}'.")
        messagebox.showinfo(
            "Unique elements",
            f"Written to sheet '{summary['sheet']}' in "
            f"{os.path.basename(summary['path'])}.\n\n"
            + "\n".join(f"{col}: {n} unique value(s)"
                        for col, n in summary["per_col"].items())
            + extra)

    # ----------------------------------------------------- Evaluation tab -- #
    def _build_ai_eval_tab(self):
        """Evaluation page: pick a stored analysis workbook + run sheet, the
        reference sections, the evaluations to compute and where to write
        them (data_extraction.evaluation.column_evaluator)."""
        ev = self._scrollable_tab("Evaluation")

        wb_wrap = ttk.LabelFrame(ev, text=" Analysis workbook (.xlsx — any sheet; "
                                          "without a 'Section' column you are "
                                          "asked which column substitutes it) ",
                                 padding=4)
        wb_wrap.pack(fill="x")
        row = ttk.Frame(wb_wrap)
        row.pack(fill="x")
        ttk.Label(row, text="Workbook:").pack(side="left")
        self.eval_wb = ttk.Entry(row)
        self.eval_wb.pack(side="left", fill="x", expand=True, padx=4)
        ttk.Button(row, text="Browse…", command=self._eval_browse_wb).pack(side="left")
        row2 = ttk.Frame(wb_wrap)
        row2.pack(fill="x", pady=(4, 0))
        ttk.Label(row2, text="Run sheet:").pack(side="left")
        self.eval_sheet = ttk.Combobox(row2, state="readonly", width=34, values=[])
        self.eval_sheet.pack(side="left", padx=4)
        ttk.Button(row2, text="Refresh sheets",
                   command=self._eval_refresh_sheets).pack(side="left")

        ref_wrap = ttk.LabelFrame(ev, text=" Reference data (the section content the "
                                           "columns were generated from) ", padding=4)
        ref_wrap.pack(fill="x", pady=(6, 0))
        jrow = ttk.Frame(ref_wrap)
        jrow.pack(fill="x")
        ttk.Label(jrow, text="Sections JSON:").pack(side="left")
        self.eval_ref_json = ttk.Entry(jrow)
        self.eval_ref_json.pack(side="left", fill="x", expand=True, padx=4)
        ttk.Button(jrow, text="Browse…",
                   command=self._eval_browse_ref_json).pack(side="left")
        ttk.Label(ref_wrap, foreground="#666666", justify="left",
                  text="Each run-sheet row is matched automatically: its 'Section' "
                       "value is looked up\namong this JSON's section titles and "
                       "evaluated against that section's text.").pack(
            anchor="w", pady=(2, 0))

        met_wrap = ttk.LabelFrame(ev, text=" Evaluations to run (each cell / item vs. "
                                           "its reference text) ", padding=4)
        met_wrap.pack(fill="x", pady=(6, 0))
        self.eval_metric_vars = {}
        self.eval_metric_btns = {}
        grid = ttk.Frame(met_wrap)
        grid.pack(fill="x")
        for i, (label, _names) in enumerate(_EVAL_METRIC_OPTIONS):
            var = tk.BooleanVar(value=True)
            self.eval_metric_vars[label] = var
            cb = ttk.Checkbutton(grid, text=label, variable=var)
            cb.grid(row=i % 4, column=i // 4, sticky="w", padx=(0, 18))
            self.eval_metric_btns[label] = cb
            if label in _EVAL_METRIC_TIPS:   # hover explanation per metric
                _Tooltip(cb, _EVAL_METRIC_TIPS[label])
        self.eval_uniq_var = tk.BooleanVar(value=True)
        uniq_cb = ttk.Checkbutton(
            met_wrap, variable=self.eval_uniq_var,
            text="Also evaluate the unique generated values "
                 "(via save_unique_elements_to_new_sheet)")
        uniq_cb.pack(anchor="w", pady=(4, 0))
        _Tooltip(uniq_cb, _EVAL_METRIC_TIPS["__uniq__"])

        # Embedding backend/service/model — only used when an embedding metric
        # (embedding cosine / BERTScore) is ticked above. Drives llm_utils'
        # get_embedding (remote) or vectorize_strings_local (local).
        emb_wrap = ttk.LabelFrame(
            ev, text=" Semantic-embedding backend (used only when an embedding "
                     "metric above is ticked) ", padding=4)
        emb_wrap.pack(fill="x", pady=(6, 0))
        emb_brow = ttk.Frame(emb_wrap)
        emb_brow.pack(fill="x")
        ttk.Label(emb_brow, text="Backend:").pack(side="left")
        self.eval_embed_backend = tk.StringVar(value="api")
        ttk.Radiobutton(emb_brow, text="Remote API", value="api",
                        variable=self.eval_embed_backend,
                        command=self._eval_backend_changed).pack(
            side="left", padx=(4, 8))
        ttk.Radiobutton(emb_brow, text="Local model", value="local",
                        variable=self.eval_embed_backend,
                        command=self._eval_backend_changed).pack(side="left")
        srow = ttk.Frame(emb_wrap)
        srow.pack(fill="x", pady=(4, 0))
        ttk.Label(srow, text="Service:").pack(side="left")
        self.eval_embed_service = ttk.Combobox(
            srow, state="readonly", width=14, values=["DLR Ollama", "BlaBla"])
        self.eval_embed_service.set("DLR Ollama")
        self.eval_embed_service.pack(side="left", padx=4)
        self.eval_embed_service.bind(
            "<<ComboboxSelected>>",
            lambda _e: self._eval_refresh_embed_models())
        ttk.Label(srow, text="Model:").pack(side="left", padx=(8, 0))
        self.eval_embed_model = ttk.Combobox(srow, width=30, values=[])
        self.eval_embed_model.pack(side="left", fill="x", expand=True, padx=4)
        ttk.Button(srow, text="List models",
                   command=self._eval_refresh_embed_models).pack(side="left")
        ttk.Label(emb_wrap, foreground="#666666", justify="left",
                  text="The divided reference sentences and candidate items are "
                       "embedded in batches and compared by cosine similarity. "
                       "Leave Model blank to use the service's default embedding "
                       "model.").pack(anchor="w", pady=(2, 0))
        self._eval_bertscore_note = ttk.Label(
            emb_wrap, foreground="#666666", justify="left", text="")
        self._eval_bertscore_note.pack(anchor="w")
        self._eval_sync_bertscore_state()

        out_wrap = ttk.LabelFrame(ev, text=" Write results to ", padding=4)
        out_wrap.pack(fill="x", pady=(6, 0))
        self.eval_out_choice = tk.StringVar(value="same")
        ttk.Radiobutton(out_wrap, text="New sheets in the same workbook "
                                       "(Eval / Uniq Eval next to the run)",
                        variable=self.eval_out_choice, value="same").pack(anchor="w")
        ttk.Radiobutton(out_wrap, text="A new evaluation workbook (asks where; also "
                                       "gets a copy of the run sheet)",
                        variable=self.eval_out_choice, value="new").pack(anchor="w")

        brow = ttk.Frame(ev)
        brow.pack(fill="x", pady=6)
        self.eval_run_btn = ttk.Button(brow, text="Run evaluation",
                                       command=self._ai_eval_run)
        self.eval_run_btn.pack(side="left")

    def _eval_browse_wb(self):
        path = filedialog.askopenfilename(
            title="Select a column-analysis workbook",
            filetypes=[("Excel workbook", "*.xlsx")])
        if path:
            self.eval_wb.delete(0, tk.END)
            self.eval_wb.insert(0, path)
            self._eval_refresh_sheets()

    def _eval_refresh_sheets(self):
        """List EVERY sheet of the workbook — exactly like the Unique
        elements tab: 'Run N …' snapshots, hand-made sheets and the
        pipeline's own result sheets alike; with or without a 'Section'
        column (a substitute is asked for at run time). Only the 'All
        sheets' bulk option skips the result sheets so an evaluation never
        evaluates its own outputs."""
        path = self.eval_wb.get().strip()
        sheets = []
        if path and os.path.exists(path):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(path, read_only=True)
                sheets = list(wb.sheetnames)
                wb.close()
            except Exception as exc:  # noqa: BLE001
                messagebox.showerror("Workbook", f"Could not read sheets:\n{exc}")
        values = (["All sheets"] + sheets) if sheets else []
        self.eval_sheet.configure(values=values)
        self.eval_sheet.set(sheets[0] if sheets else "")
        if path and not sheets:
            self.status_var.set("No sheet found in the selected workbook.")

    def _eval_ask_section_column(self, sheet, columns):
        """Modal picker shown when a sheet chosen for evaluation has no
        'Section' column: choose the column that substitutes it — its values
        become the per-row key matched against the reference section titles;
        every other column is evaluated. Returns the column name, or None
        when cancelled."""
        dlg = tk.Toplevel(self.root)
        dlg.title(f"No 'Section' column in '{sheet}'")
        dlg.transient(self.root.winfo_toplevel())
        frm = ttk.Frame(dlg, padding=12)
        frm.pack(fill="both", expand=True)
        ttk.Label(frm, justify="left",
                  text=f"Sheet '{sheet}' has no 'Section' column.\n\n"
                       "Choose the column that substitutes it: its values are "
                       "the per-row key\nmatched against the reference section "
                       "titles. All other columns are\nevaluated against the "
                       "matched section text.").pack(anchor="w")
        combo = ttk.Combobox(frm, state="readonly", values=list(columns),
                             width=44)
        # Preselect a likely key column by name, else the first one.
        guess = next((c for c in columns
                      if any(h in str(c).lower()
                             for h in ("section", "title", "rule", "name", "id"))),
                     columns[0])
        combo.set(guess)
        combo.pack(fill="x", pady=8)
        out = {"col": None}
        btns = ttk.Frame(frm)
        btns.pack(fill="x")

        def ok():
            out["col"] = combo.get() or None
            dlg.destroy()

        ttk.Button(btns, text="Use this column", command=ok).pack(side="left")
        ttk.Button(btns, text="Cancel",
                   command=dlg.destroy).pack(side="left", padx=6)
        dlg.bind("<Return>", lambda _e: ok())
        dlg.bind("<Escape>", lambda _e: dlg.destroy())
        dlg.grab_set()
        self.root.wait_window(dlg)
        return out["col"]

    def _eval_browse_ref_json(self):
        path = filedialog.askopenfilename(
            title="Select the sections JSON the analysis was made from",
            filetypes=[("JSON files", "*.json")])
        if path:
            self.eval_ref_json.delete(0, tk.END)
            self.eval_ref_json.insert(0, path)

    def _eval_selected_metrics(self):
        metrics = []
        for label, names in _EVAL_METRIC_OPTIONS:
            if self.eval_metric_vars[label].get():
                metrics.extend(names)
        return metrics

    def _eval_backend_changed(self):
        """Embedding backend switched: BERTScore is local-only, so keep its
        checkbox in step before refreshing the model list."""
        self._eval_sync_bertscore_state()
        self._eval_refresh_embed_models()

    def _eval_sync_bertscore_state(self):
        """Enable the BERTScore option only on the local backend.

        BERTScore needs one vector per TOKEN. The remote /embeddings endpoint
        returns a single pooled vector per string, and each evaluation scores one
        sentence against one item — so the cosine matrix would be 1x1 and P, R
        and F1 would each collapse to the pooled cosine, i.e. three columns
        identical to embedding_cosine. Rather than write those look-alike
        numbers, the option is unticked and greyed out on the API backend (and
        column_evaluator records None if it is forced on some other way)."""
        btn = getattr(self, "eval_metric_btns", {}).get(_BERTSCORE_LABEL)
        note = getattr(self, "_eval_bertscore_note", None)
        if btn is None:
            return
        if self.eval_embed_backend.get() == "local":
            btn.configure(state="normal")
            if note is not None:
                note.configure(
                    text="BERTScore: true token-level P/R/F1 from the local "
                         "model.")
        else:
            self.eval_metric_vars[_BERTSCORE_LABEL].set(False)
            btn.configure(state="disabled")
            if note is not None:
                note.configure(
                    text="BERTScore needs token-level embeddings → local backend "
                         "only. The remote API returns one pooled vector per "
                         "string, so P/R/F1 could only repeat the cosine.")

    def _embedding_config_for(self, metrics):
        """Embedding config dict for column_evaluator (backend/service/model +
        the injected llm_utils module), or None when no embedding metric is in
        ``metrics`` so non-embedding runs stay untouched."""
        from data_extraction.evaluation.column_evaluator import _EMBEDDING as EMB
        if not any(m in EMB for m in metrics):
            return None
        return {
            "enabled": True,
            "backend": self.eval_embed_backend.get(),
            "service": (self.eval_embed_service.get().strip() or None),
            "model": (self.eval_embed_model.get().strip() or None),
            "llm_utils": _llm_utils_mod,
        }

    def _eval_embedding_config(self):
        """Embedding config for the metrics currently ticked on the tab."""
        return self._embedding_config_for(self._eval_selected_metrics())

    def _eval_refresh_embed_models(self):
        """List the chosen service's embedding models into the Model combobox
        (API backend only; the local backend uses llm_utils' local model)."""
        if self.eval_embed_backend.get() != "api":
            self.eval_embed_model.configure(values=[])
            self.status_var.set("Local backend: embeddings come from llm_utils' "
                                "configured local model (Model field ignored).")
            return
        service = self.eval_embed_service.get().strip()
        if not service:
            return
        self.status_var.set(f"Listing {service} embedding models…")

        def work():
            try:
                models = list_embedding_models(service)
                default = get_default_embedding_model(service) if models else ""
            except Exception as exc:  # noqa: BLE001
                self.root.after(0, lambda e=exc: self.status_var.set(
                    f"Could not list embedding models: {e}"))
                return

            def apply():
                self.eval_embed_model.configure(values=models)
                if default and not self.eval_embed_model.get().strip():
                    self.eval_embed_model.set(default)
                self.status_var.set(
                    f"{len(models)} {service} embedding model(s) available"
                    if models else f"No embedding models found for {service}.")
            self.root.after(0, apply)

        threading.Thread(target=work, daemon=True).start()

    def _eval_references(self):
        """Load {section title: text} from the chosen sections JSON — the
        run-sheet 'Section' values are matched against these titles
        automatically. Returns None (after telling the user) on error."""
        path = self.eval_ref_json.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showinfo(
                "Sections JSON needed",
                "Pick the sections JSON the analysis was made from — the "
                "'Section' column of the run sheet is matched against its "
                "section titles to find each reference text.")
            return None
        try:
            from data_extraction.evaluation.column_evaluator import references_from_json
            refs = references_from_json(path)
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("Sections JSON",
                                 f"Could not load sections:\n{exc}")
            return None
        if not refs:
            messagebox.showinfo("Sections JSON",
                                f"No sections found in {os.path.basename(path)}.")
            return None
        return refs

    def _ai_eval_run(self):
        """Run the selected evaluations on the chosen workbook/run sheet(s),
        against the sections JSON. Standalone: works on any workbook that
        already holds analyzed data, independent of the Column analysis tab."""
        if self._eval_busy or self._ai_busy:
            messagebox.showinfo("Busy", "Another run is already in progress.")
            return
        path = self.eval_wb.get().strip()
        if not path or os.path.splitext(path)[1].lower() != ".xlsx" \
                or not os.path.exists(path):
            messagebox.showinfo("No workbook",
                                "Pick an .xlsx with analyzed data first — any "
                                "data sheet can be evaluated (Browse… to an "
                                "existing workbook).")
            return
        sheet = self.eval_sheet.get().strip()
        run_sheets = (None if not sheet or sheet in ("All sheets", "All runs")
                      else [sheet])
        metrics = self._eval_selected_metrics()
        if not metrics:
            messagebox.showinfo("No evaluations",
                                "Tick at least one evaluation to run.")
            return
        embedding = self._eval_embedding_config()
        self._ui_state_save()   # remember workbook + metric choices
        references = self._eval_references()
        if references is None:
            return
        # Resolve the sheet list now: sheets WITHOUT a 'Section' column need
        # the user to choose which column substitutes it (the per-row key
        # matched against the reference section titles).
        try:
            from data_extraction.evaluation.column_evaluator import (
                evaluatable_sheets, sheet_columns)
            sheets = (run_sheets if run_sheets
                      else evaluatable_sheets(path, require_section=False))
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("Workbook", f"Could not read sheets:\n{exc}")
            return
        if not sheets:
            messagebox.showinfo("Evaluation",
                                "The workbook has no evaluatable data sheet.")
            return
        section_columns = {}
        for s in sheets:
            cols = sheet_columns(path, s)
            if "Section" in cols:
                continue
            if len(cols) < 2:
                messagebox.showerror(
                    "Evaluation",
                    f"Sheet '{s}' has {len(cols)} column(s) — it needs a "
                    "section-key column plus at least one data column.")
                return
            pick = self._eval_ask_section_column(s, cols)
            if not pick:
                self.status_var.set("Evaluation cancelled — no substitute "
                                    f"for the 'Section' column of '{s}' chosen.")
                return
            section_columns[s] = pick
        run_sheets = sheets
        out_path = None
        if self.eval_out_choice.get() == "new":
            out_path = filedialog.asksaveasfilename(
                title="Write the evaluation workbook to…",
                defaultextension=".xlsx",
                filetypes=[("Excel workbook", "*.xlsx")],
                initialfile=Path(path).stem + "_evaluation.xlsx",
                confirmoverwrite=False)
            if not out_path:
                self.status_var.set("Evaluation cancelled — no output file chosen.")
                return
        uniq = "all" if self.eval_uniq_var.get() else None

        self._eval_busy = True
        self.eval_run_btn.configure(state="disabled")
        self.ai_progress.configure(mode="indeterminate")
        self.ai_progress.start(80)
        self._ai_log(f"=== Evaluation: {os.path.basename(path)} "
                     f"[{sheet or 'all runs'}], {len(references)} reference "
                     f"section(s) from {os.path.basename(self.eval_ref_json.get().strip())}, "
                     f"metrics [{', '.join(metrics)}]"
                     + (", uniques on" if uniq else "")
                     + (f", output → {os.path.basename(out_path)}" if out_path else "")
                     + " ===")
        for s, c in section_columns.items():
            self._ai_log(f"   '{s}' has no 'Section' column — using '{c}' "
                         "as the section key")
        self.status_var.set(f"Evaluating {sheet or 'all runs'} against "
                            f"{len(references)} reference section(s)…")
        if embedding:
            self._ai_log(
                f"   semantic embeddings: backend={embedding['backend']}"
                + (f", service={embedding['service']}"
                   if embedding.get('service') else "")
                + (f", model={embedding['model']}"
                   if embedding.get('model') else " (service default model)"))

        def work():
            try:
                from data_extraction.evaluation.column_evaluator import evaluate_workbook
                results = evaluate_workbook(path, references, metrics=metrics,
                                            run_sheets=run_sheets,
                                            uniq_columns=uniq, out_path=out_path,
                                            log=self._ai_log_bg,
                                            embedding=embedding,
                                            section_columns=section_columns or None)
            except Exception as exc:  # noqa: BLE001
                self.root.after(0, lambda e=exc: self._ai_eval_done(None, e))
            else:
                self.root.after(0, lambda r=results: self._ai_eval_done(r, None))

        threading.Thread(target=work, daemon=True).start()

    def _ai_eval_done(self, results, exc):
        self._eval_busy = False
        self.eval_run_btn.configure(state="normal")
        self._ai_progress_done()
        if exc is not None:
            self.status_var.set(f"Evaluation failed: {exc}")
            self._ai_log(f"[ERROR] evaluation failed: {exc}")
            messagebox.showerror("Evaluation failed", str(exc))
            return
        if not results:
            self.status_var.set("Evaluation found no data sheet to evaluate.")
            messagebox.showinfo("Evaluation",
                                "The workbook has no evaluatable data sheet.")
            return
        lines = []
        for sheet, s in results.items():
            line = f"{sheet}: {s['rows']} row(s)"
            if s.get("coverage") is not None:
                line += (f", grounded {s['true']}/{s['true'] + s['false']}"
                         f" ({s['coverage']}%)")
            line += f" → '{s['eval_sheet']}'"
            if s.get("uniq_eval_sheet"):
                line += f", uniques → '{s['uniq_eval_sheet']}'"
            lines.append(line)
        first = next(iter(results.values()))
        where = os.path.basename(first["out_path"])
        self._ai_note_output(first["out_path"])
        self.status_var.set(f"Evaluation complete → {where}: " + " | ".join(lines))
        self._ai_log(f"=== Evaluation complete → {first['out_path']} ===")
        for line in lines:
            self._ai_log("   " + line)
        messagebox.showinfo("Evaluation complete",
                            f"Written to {first['out_path']}\n\n" + "\n".join(lines))

    def _col_ask_auto_eval(self):
        """Modal picker shown when 'Analyze queued' starts with the
        auto-evaluate box ticked: choose which of the possible evaluations
        run automatically on this batch once the analysis finishes. The
        checkboxes are shared with the Evaluation tab. Returns True to
        evaluate afterwards, False for 'Skip evaluation'."""
        dlg = tk.Toplevel(self.root)
        dlg.title("Auto-evaluation of this analysis")
        dlg.transient(self.root.winfo_toplevel())
        frm = ttk.Frame(dlg, padding=12)
        frm.pack(fill="both", expand=True)
        ttk.Label(frm, justify="left",
                  text="Choose the evaluations to run on each section result "
                       "as soon as it is saved\n(each cell / item is compared "
                       "with the section text it was generated from):").pack(anchor="w")
        # The picker keeps its own choices (self._col_eval_choices) — fully
        # independent of the Evaluation tab.
        dlg_vars = {}
        for label, _names in _EVAL_METRIC_OPTIONS:
            var = tk.BooleanVar(value=self._col_eval_choices.get(label, True))
            dlg_vars[label] = var
            ttk.Checkbutton(frm, text=label, variable=var).pack(anchor="w", padx=10)
        uniq_var = tk.BooleanVar(value=self._col_eval_choices.get("__uniq__", True))
        ttk.Checkbutton(frm, variable=uniq_var,
                        text="Also evaluate the unique generated values "
                             "(element + count, grounded in the references)").pack(
            anchor="w", pady=(8, 0))
        choice = {"ok": False}

        def _ok():
            if not any(v.get() for v in dlg_vars.values()):
                messagebox.showinfo(
                    "No evaluations",
                    "Tick at least one evaluation — or press Skip evaluation.",
                    parent=dlg)
                return
            self._col_eval_choices = {label: v.get() for label, v in dlg_vars.items()}
            self._col_eval_choices["__uniq__"] = uniq_var.get()
            choice["ok"] = True
            dlg.destroy()

        btns = ttk.Frame(frm)
        btns.pack(fill="x", pady=(12, 0))
        ttk.Button(btns, text="OK — evaluate after the analysis",
                   command=_ok).pack(side="left")
        ttk.Button(btns, text="Skip evaluation",
                   command=dlg.destroy).pack(side="right")
        dlg.grab_set()
        self.root.wait_window(dlg)
        return choice["ok"]

    def _col_eval_saved_row(self, row):
        """Evaluate one just-saved section result (the chosen evaluations vs.
        its reference text) and refresh the 'Eval <run sheet>' — and, when
        chosen, the 'Uniq Eval <run sheet>' — before the next section is
        analyzed. Failures only reach the status bar; the analysis goes on."""
        if not (self._col_run_auto_eval and self._col_run_eval_metrics
                and self._col_run_sheet):
            return
        try:
            from data_extraction.evaluation.column_evaluator import (
                _lookup_reference, evaluate_row, evaluate_uniques,
                write_eval_sheet, configure_embeddings)
            # If an embedding metric is part of this auto-eval run, configure the
            # backend (from the Evaluation tab's widgets) so evaluate_row can
            # compute it; harmless/no-op when none is selected.
            try:
                emb_cfg = self._embedding_config_for(self._col_run_eval_metrics)
                if emb_cfg:
                    configure_embeddings(**emb_cfg)
            except Exception:  # noqa: BLE001 - never block the analysis
                pass
            data_cols = [c for c in self._col_table_cols if c != "Section"]
            title = str(row.get("Section", ""))
            reference = _lookup_reference(self._col_run_refs, title)
            self._ai_log(f"⚖ Evaluating '{title}' with "
                         f"[{', '.join(self._col_run_eval_metrics)}]")
            self._ai_log(f"   reference ({len(reference)} chars): "
                         f"{self._ai_preview(reference)}")
            for col in data_cols:
                self._ai_log(f"   evaluated text [{col}]: "
                             f"{self._ai_preview(row.get(col))}")
            entry = evaluate_row(row, self._col_run_refs, data_cols,
                                 self._col_run_eval_metrics)
            if self._col_rerun:
                # replace the failed row's earlier evaluation
                for i, e in enumerate(self._col_eval_entries):
                    if e.get("Section") == entry.get("Section"):
                        self._col_eval_entries[i] = entry
                        break
                else:
                    self._col_eval_entries.append(entry)
            else:
                self._col_eval_entries.append(entry)
            self._ai_log("   result: " + self._ai_preview(
                "; ".join(f"{k}={v}" for k, v in entry.items()
                          if k not in ("Section",)), 300))
            write_eval_sheet(self._col_store_path, self._col_run_sheet,
                             self._col_eval_entries)
            self._ai_log(f"   → 'Eval {self._col_run_sheet}' refreshed"[:120])
            if self._col_run_eval_uniq:
                uniq_cols = [c for c in self._col_run_uniq_cols if c != "Section"]
                if uniq_cols:
                    evaluate_uniques(self._col_store_path, self._col_run_sheet,
                                     self._col_run_refs, uniq_cols)
                    self._ai_log(f"   → 'Uniq Eval {self._col_run_sheet}' "
                                 f"refreshed ({', '.join(uniq_cols)})"[:150])
        except Exception as exc:  # noqa: BLE001 - never interrupt the analysis
            self.status_var.set(f"Row evaluation failed: {exc}")
            self._ai_log(f"[ERROR] row evaluation failed: {exc}")

    def _col_eval_finish(self):
        """After the batch: when rows were evaluated along the way, show the
        overall result in the status bar."""
        sheet = self._col_run_sheet
        if not (sheet and self._col_run_auto_eval and self._col_eval_entries):
            return
        parts = [f"{self.status_var.get()} Evaluated row by row → "
                 f"'Eval {sheet}'"[:200]]
        if "grounding" in self._col_run_eval_metrics:
            t = sum(e.get("Row True") or 0 for e in self._col_eval_entries)
            f = sum(e.get("Row False") or 0 for e in self._col_eval_entries)
            if t + f:
                parts.append(f"grounded {t}/{t + f} "
                             f"({round(100.0 * t / (t + f), 1)}%)")
        self.status_var.set(", ".join(parts) + ".")
        self._ai_log(f"=== Evaluation of the run finished: "
                     f"{len(self._col_eval_entries)} section(s) evaluated"
                     + (f", {parts[-1]}" if len(parts) > 1 else "") + " ===")

    def _col_export(self):
        if not self._col_results:
            messagebox.showinfo("No table", "Run a column analysis first, then export.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv"), ("JSON", "*.json")],
            initialfile=self._default_export_name("ai_columns", "xlsx"),
            confirmoverwrite=False)
        if not path:
            return
        try:
            sheet = self._col_write_store(path, self._col_table_cols, self._col_results)
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("Export failed", str(exc))
            return
        where = path + (f" (sheet '{sheet}')" if sheet else "")
        self.status_var.set(f"Exported {len(self._col_results)} analysis row(s) → {where}")
        messagebox.showinfo("Export complete",
                            f"Wrote {len(self._col_results)} row(s) to:\n{where}")