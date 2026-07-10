"""
Interactive review UI for the EASA structured-extraction JSON.

Opens the JSON produced by EASA_Parser (``{document_metadata, rules_hierarchy}``)
and lets you navigate the regulation hierarchy as an expandable tree, with a
details pane per node: full text, EASA attributes, hyperlinks, and extracted
images/tables. Images preview inline (Pillow when available for JPG/BMP/TIFF/…,
Tk's built-in PNG/GIF otherwise); tables (.xlsx) preview as a grid via openpyxl.
A live search filters the tree, a Summary reviews the whole document at a glance,
and the Export menu writes a node index (CSV/Excel), the full text (Markdown),
or the selected subtree (JSON).

Run standalone:
    python -m data_extraction.easa.json_review_ui [path/to/structured.json]

Also embeddable: pass a container Frame as ``root`` (the Data Extraction Studio
hosts it this way).
"""

import json
import os
import subprocess
import sys
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, scrolledtext, ttk

from data_extraction.ai_utils.review_panel import AIReviewMixin


def _open_externally(path: str):
    """Open a file with the OS default application (cross-platform)."""
    try:
        if sys.platform.startswith("win"):
            os.startfile(path)  # noqa: SIM115 - Windows API
        elif sys.platform == "darwin":
            subprocess.Popen(["open", path])
        else:
            subprocess.Popen(["xdg-open", path])
    except Exception as exc:  # noqa: BLE001
        messagebox.showerror("Could not open file", f"{path}\n\n{exc}")


class EASAJsonReviewApp(AIReviewMixin):
    def __init__(self, root, json_path: str = None, logger=None):
        self.root = root
        self.logger = logger
        if isinstance(root, (tk.Tk, tk.Toplevel)):
            root.title("EASA Structured JSON — Review")
            root.geometry("1280x820")
            root.minsize(960, 600)

        self.data = None
        self.json_path = None
        self.assets_base = None       # dir that holds images/ and tables/
        self.node_by_iid = {}
        self._node_count = 0
        self._checked = {}            # id(node) -> node, survives re-filtering
        self._preview_img = None      # keep a ref so Tk doesn't GC the preview
        self._current_node = None     # node shown in the details pane
        self._init_ai_state()         # AI Review page state (AIReviewMixin)

        self._build_ui()

        if json_path:
            self._load(json_path)

    # ------------------------------------------------------------------ UI -- #
    def _build_ui(self):
        # Top toolbar: file + search
        bar = ttk.Frame(self.root, padding=(8, 6))
        bar.pack(fill="x")

        ttk.Label(bar, text="Structured JSON:").pack(side="left")
        self.ent_path = ttk.Entry(bar, width=60)
        self.ent_path.pack(side="left", padx=5)
        ttk.Button(bar, text="Browse…", command=self._browse).pack(side="left")
        ttk.Button(bar, text="Load", command=self._load_from_entry).pack(side="left", padx=(4, 12))

        ttk.Label(bar, text="Search:").pack(side="left")
        self.ent_search = ttk.Entry(bar, width=26)
        self.ent_search.pack(side="left", padx=5)
        self.ent_search.bind("<Return>", lambda e: self._apply_filter())
        ttk.Button(bar, text="Find", command=self._apply_filter).pack(side="left")
        ttk.Button(bar, text="Clear", command=self._clear_filter).pack(side="left", padx=(4, 0))

        # Right-aligned: review + export actions
        self._build_export_menu(bar)
        ttk.Button(bar, text="Summary", command=self._show_summary).pack(side="right", padx=(0, 6))

        # Document metadata line
        self.meta_var = tk.StringVar(value="No file loaded.")
        ttk.Label(self.root, textvariable=self.meta_var, foreground="#444444",
                  padding=(10, 0)).pack(fill="x")

        # Top-level pages: Browse & Review | AI Review
        self.main_nb = ttk.Notebook(self.root)
        self.main_nb.pack(fill="both", expand=True, padx=8, pady=6)

        browse = ttk.Frame(self.main_nb)
        self.main_nb.add(browse, text="Browse & Review")

        # Main split: tree (left) | details (right)
        paned = ttk.PanedWindow(browse, orient="horizontal")
        paned.pack(fill="both", expand=True)

        left = ttk.Frame(paned)
        paned.add(left, weight=1)

        tree_tools = ttk.Frame(left)
        tree_tools.pack(fill="x")
        ttk.Button(tree_tools, text="Expand all", command=lambda: self._set_all_open(True)).pack(side="left")
        ttk.Button(tree_tools, text="Collapse all", command=lambda: self._set_all_open(False)).pack(side="left", padx=4)
        ttk.Button(tree_tools, text="Select all", command=self._select_all_checks).pack(side="left", padx=(8, 0))
        ttk.Button(tree_tools, text="Clear checks", command=self._clear_checks).pack(side="left", padx=4)
        self.check_var = tk.StringVar(value="0 checked")
        ttk.Label(tree_tools, textvariable=self.check_var, foreground="#444444").pack(side="right")

        self.tree = ttk.Treeview(left, columns=("sel",), show="tree headings", selectmode="browse")
        self.tree.heading("#0", text="Node")
        self.tree.heading("sel", text="✓", command=self._toggle_check_all)
        self.tree.column("sel", width=34, minwidth=28, anchor="center", stretch=False)
        yscroll = ttk.Scrollbar(left, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=yscroll.set)
        yscroll.pack(side="right", fill="y")
        self.tree.pack(side="left", fill="both", expand=True, pady=(4, 0))
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.tree.bind("<Button-1>", self._on_tree_click)
        self.tree.bind("<space>", self._toggle_check_selected)

        right = ttk.Frame(paned)
        paned.add(right, weight=2)

        self.header_var = tk.StringVar(value="Select a node to see its details.")
        ttk.Label(right, textvariable=self.header_var, font=("TkDefaultFont", 11, "bold"),
                  wraplength=680, justify="left").pack(anchor="w", pady=(0, 4))

        self.detail_nb = ttk.Notebook(right)
        self.detail_nb.pack(fill="both", expand=True)
        self._build_detail_tabs()

        # Full-width AI Review page (second top-level tab)
        self._build_ai_tab()

        # Status bar
        self.status_var = tk.StringVar(value="")
        ttk.Label(self.root, textvariable=self.status_var, relief="sunken",
                  anchor="w", padding=(6, 2)).pack(fill="x", side="bottom")

    def _build_detail_tabs(self):
        # Overview (ScrolledText must live inside a container frame, not be added
        # to the Notebook directly — its real parent is an internal frame).
        ov = ttk.Frame(self.detail_nb)
        self.tab_overview = scrolledtext.ScrolledText(ov, wrap="word", height=8)
        self.tab_overview.configure(state="disabled")
        self.tab_overview.pack(fill="both", expand=True)
        self.detail_nb.add(ov, text="Overview")

        # Text
        tx = ttk.Frame(self.detail_nb)
        self.tab_text = scrolledtext.ScrolledText(tx, wrap="word")
        self.tab_text.configure(state="disabled")
        self.tab_text.pack(fill="both", expand=True)
        self.detail_nb.add(tx, text="Text")

        # Attributes
        attr_frame = ttk.Frame(self.detail_nb)
        self.tab_attrs = ttk.Treeview(attr_frame, columns=("value",), show="tree headings")
        self.tab_attrs.heading("#0", text="Attribute")
        self.tab_attrs.heading("value", text="Value")
        self.tab_attrs.column("#0", width=220, anchor="w")
        self.tab_attrs.column("value", width=440, anchor="w")
        a_scroll = ttk.Scrollbar(attr_frame, orient="vertical", command=self.tab_attrs.yview)
        self.tab_attrs.configure(yscrollcommand=a_scroll.set)
        a_scroll.pack(side="right", fill="y")
        self.tab_attrs.pack(side="left", fill="both", expand=True)
        self.detail_nb.add(attr_frame, text="Attributes")

        # Hyperlinks
        link_frame = ttk.Frame(self.detail_nb)
        self.tab_links = ttk.Treeview(link_frame, columns=("target",), show="tree headings")
        self.tab_links.heading("#0", text="Link text")
        self.tab_links.heading("target", text="Target")
        self.tab_links.column("#0", width=300, anchor="w")
        self.tab_links.column("target", width=360, anchor="w")
        l_scroll = ttk.Scrollbar(link_frame, orient="vertical", command=self.tab_links.yview)
        self.tab_links.configure(yscrollcommand=l_scroll.set)
        l_scroll.pack(side="right", fill="y")
        self.tab_links.pack(side="left", fill="both", expand=True)
        self.detail_nb.add(link_frame, text="Hyperlinks")

        # Assets (images + tables) with a switchable preview pane
        assets = ttk.Frame(self.detail_nb)

        lists = ttk.Frame(assets)
        lists.pack(side="left", fill="y")

        ttk.Label(lists, text="Images:").pack(anchor="w")
        self.lst_images = tk.Listbox(lists, height=8, width=36, exportselection=False)
        self.lst_images.pack(fill="x")
        self.lst_images.bind("<<ListboxSelect>>", self._on_image_select)
        self.lst_images.bind("<Double-Button-1>", lambda e: self._open_selected_asset(self.lst_images, "images"))

        ttk.Label(lists, text="Tables:").pack(anchor="w", pady=(6, 0))
        self.lst_tables = tk.Listbox(lists, height=6, width=36, exportselection=False)
        self.lst_tables.pack(fill="x")
        self.lst_tables.bind("<<ListboxSelect>>", self._on_table_select)
        self.lst_tables.bind("<Double-Button-1>", lambda e: self._open_selected_asset(self.lst_tables, "tables"))

        ttk.Label(lists, text="(double-click to open externally)", foreground="#777").pack(
            anchor="w", pady=(6, 0))

        # Preview pane holds an image/message label OR a table grid; only one
        # is packed at a time via _show_preview_widget().
        self.preview = ttk.LabelFrame(assets, text=" Preview ", padding=6)
        self.preview.pack(side="right", fill="both", expand=True, padx=(8, 0))

        self.preview_label = ttk.Label(self.preview, anchor="center",
                                       text="Select an image or table to preview.")

        self.table_wrap = ttk.Frame(self.preview)
        self.table_preview = ttk.Treeview(self.table_wrap, show="headings", height=12)
        tp_y = ttk.Scrollbar(self.table_wrap, orient="vertical", command=self.table_preview.yview)
        tp_x = ttk.Scrollbar(self.table_wrap, orient="horizontal", command=self.table_preview.xview)
        self.table_preview.configure(yscrollcommand=tp_y.set, xscrollcommand=tp_x.set)
        tp_y.pack(side="right", fill="y")
        tp_x.pack(side="bottom", fill="x")
        self.table_preview.pack(side="left", fill="both", expand=True)

        self._show_preview_widget("message")

        self.detail_nb.add(assets, text="Images & Tables")

    def _show_preview_widget(self, kind):
        """Show the image/message label ('image'/'message') or the table grid ('table')."""
        if kind == "table":
            self.preview_label.pack_forget()
            self.table_wrap.pack(fill="both", expand=True)
        else:
            self.table_wrap.pack_forget()
            self.preview_label.pack(fill="both", expand=True)

    # --------------------------------------------------------------- load -- #
    def _browse(self):
        path = filedialog.askopenfilename(
            title="Select EASA structured JSON",
            filetypes=[("JSON files", "*.json")],
        )
        if path:
            self.ent_path.delete(0, tk.END)
            self.ent_path.insert(0, path)
            self._load(path)

    def _load_from_entry(self):
        path = self.ent_path.get().strip()
        if path:
            self._load(path)

    def _load(self, path):
        if not os.path.exists(path):
            messagebox.showerror("Not found", f"File does not exist:\n{path}")
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except (OSError, json.JSONDecodeError) as exc:
            messagebox.showerror("Load failed", f"Could not read JSON:\n{exc}")
            return

        # Accept {document_metadata, rules_hierarchy} or a bare hierarchy list.
        if isinstance(data, dict):
            hierarchy = data.get("rules_hierarchy")
            metadata = data.get("document_metadata", {})
        elif isinstance(data, list):
            hierarchy = data
            metadata = {}
        else:
            hierarchy = None
            metadata = {}

        if not hierarchy:
            messagebox.showerror(
                "Unexpected format",
                "This file has no 'rules_hierarchy'. It does not look like an EASA "
                "structured extraction JSON.",
            )
            return

        self.data = data
        self.json_path = path
        self.assets_base = os.path.dirname(os.path.abspath(path))
        self.ent_path.delete(0, tk.END)
        self.ent_path.insert(0, path)

        title = metadata.get("source-title") or Path(path).stem
        domain = metadata.get("Domain", "")
        self.meta_var.set(f"📄 {title}" + (f"   ·   Domain: {domain}" if domain else ""))

        self._hierarchy = hierarchy
        self._checked.clear()
        self._update_check_status()
        self._populate_tree()

    # --------------------------------------------------------------- tree -- #
    def _node_label(self, node):
        attrs = node.get("attributes", {}) or {}
        title = attrs.get("source-title") or attrs.get("title")
        etype = node.get("element_type", "node")
        badges = ""
        if node.get("extracted_images"):
            badges += f"  🖼{len(node['extracted_images'])}"
        if node.get("extracted_tables"):
            badges += f"  ▦{len(node['extracted_tables'])}"
        if title:
            return f"{title}  [{etype}]{badges}"
        sid = attrs.get("sdt-id", "")
        return f"{etype} ({sid}){badges}"

    def _subtree_matches(self, node, filt):
        if filt in self._node_label(node).lower():
            return True
        if filt in (node.get("text_content", "") or "").lower():
            return True
        return any(self._subtree_matches(c, filt) for c in (node.get("children") or []))

    def _insert_node(self, parent_iid, node, filt):
        if filt and not self._subtree_matches(node, filt):
            return
        iid = self.tree.insert(parent_iid, "end", text=self._node_label(node),
                               values=("☑" if id(node) in self._checked else "☐",))
        self.node_by_iid[iid] = node
        self._node_count += 1
        for child in (node.get("children") or []):
            self._insert_node(iid, child, filt)
        if filt:
            self.tree.item(iid, open=True)

    def _populate_tree(self, filt=None):
        self.tree.delete(*self.tree.get_children())
        self.node_by_iid.clear()
        self._node_count = 0
        for node in self._hierarchy:
            self._insert_node("", node, filt)

        if filt:
            self.status_var.set(f"Filter '{filt}': {self._node_count} matching node(s).")
        else:
            self.status_var.set(f"Loaded {self._node_count} node(s).")
            # Open the first level for orientation.
            for iid in self.tree.get_children():
                self.tree.item(iid, open=True)

    def _apply_filter(self):
        if not getattr(self, "_hierarchy", None):
            return
        filt = self.ent_search.get().strip().lower()
        self._populate_tree(filt or None)

    def _clear_filter(self):
        self.ent_search.delete(0, tk.END)
        if getattr(self, "_hierarchy", None):
            self._populate_tree(None)

    def _set_all_open(self, is_open):
        def walk(iid):
            self.tree.item(iid, open=is_open)
            for c in self.tree.get_children(iid):
                walk(c)
        for iid in self.tree.get_children():
            walk(iid)

    # -------------------------------------------------------- checkboxes -- #
    def _set_check(self, iid, node, checked):
        if checked:
            self._checked[id(node)] = node
        else:
            self._checked.pop(id(node), None)
        self.tree.set(iid, "sel", "☑" if checked else "☐")

    def _on_tree_click(self, event):
        # Toggle only when the click lands in the ✓ column; everything else
        # keeps the normal select/expand behaviour.
        if self.tree.identify("region", event.x, event.y) != "cell":
            return None
        if self.tree.identify_column(event.x) != "#1":
            return None
        iid = self.tree.identify_row(event.y)
        node = self.node_by_iid.get(iid)
        if node is None:
            return None
        self._set_check(iid, node, id(node) not in self._checked)
        self._update_check_status()
        return "break"

    def _toggle_check_selected(self, event=None):
        for iid in self.tree.selection():
            node = self.node_by_iid.get(iid)
            if node is not None:
                self._set_check(iid, node, id(node) not in self._checked)
        self._update_check_status()
        return "break"

    def _select_all_checks(self):
        """Check every visible node (respects an active search filter)."""
        for iid, node in self.node_by_iid.items():
            self._set_check(iid, node, True)
        self._update_check_status()

    def _clear_checks(self):
        self._checked.clear()
        for iid in self.node_by_iid:
            self.tree.set(iid, "sel", "☐")
        self._update_check_status()

    def _toggle_check_all(self):
        """✓ heading click: select all visible, or clear if all are checked."""
        if self.node_by_iid and all(id(n) in self._checked
                                    for n in self.node_by_iid.values()):
            self._clear_checks()
        else:
            self._select_all_checks()

    def _update_check_status(self):
        self.check_var.set(f"{len(self._checked)} checked")

    def _checked_in_doc_order(self):
        """Checked nodes in document order (including any hidden by a filter)."""
        out = []

        def walk(node):
            if id(node) in self._checked:
                out.append(node)
            for child in (node.get("children") or []):
                walk(child)

        for node in (getattr(self, "_hierarchy", None) or []):
            walk(node)
        return out

    # ------------------------------------------------------------ details -- #
    def _set_text(self, widget, content):
        widget.configure(state="normal")
        widget.delete("1.0", tk.END)
        widget.insert(tk.END, content or "")
        widget.configure(state="disabled")

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        node = self.node_by_iid.get(sel[0])
        if node is None:
            return
        self._show_node(node)

    def _show_node(self, node):
        self._current_node = node
        attrs = node.get("attributes", {}) or {}
        title = attrs.get("source-title") or attrs.get("title") or node.get("element_type", "node")
        self.header_var.set(f"{title}")

        text_content = node.get("text_content", "") or ""
        links = node.get("hyperlinks", []) or []
        images = node.get("extracted_images", []) or []
        tables = node.get("extracted_tables", []) or []

        # Overview
        overview = (
            f"Element type : {node.get('element_type', '')}\n"
            f"sdt-id       : {attrs.get('sdt-id', '')}\n"
            f"Title        : {title}\n"
            f"Children     : {len(node.get('children') or [])}\n"
            f"Hyperlinks   : {len(links)}\n"
            f"Images       : {len(images)}\n"
            f"Tables       : {len(tables)}\n"
            f"Text length  : {len(text_content)} chars\n"
        )
        self._set_text(self.tab_overview, overview)

        # Text
        self._set_text(self.tab_text, text_content)

        # Attributes
        self.tab_attrs.delete(*self.tab_attrs.get_children())
        for k, v in attrs.items():
            self.tab_attrs.insert("", "end", text=str(k), values=(str(v),))

        # Hyperlinks
        self.tab_links.delete(*self.tab_links.get_children())
        for link in links:
            self.tab_links.insert("", "end", text=str(link.get("text", "")),
                                  values=(str(link.get("target", "")),))

        # Assets
        self.lst_images.delete(0, tk.END)
        for img in images:
            self.lst_images.insert(tk.END, img)
        self.lst_tables.delete(0, tk.END)
        for tbl in tables:
            self.lst_tables.insert(tk.END, tbl)
        self._show_preview_widget("message")
        self.preview_label.configure(image="", text="Select an image or table to preview.")
        self._preview_img = None

    def _resolve_asset(self, filename, kind):
        """Resolve an image/table filename to an on-disk path next to the JSON."""
        if not self.assets_base:
            return None
        candidate = os.path.join(self.assets_base, kind, filename)
        if os.path.exists(candidate):
            return candidate
        # Fall back to the filename as-is (in case it is already a full path).
        return filename if os.path.exists(filename) else None

    def _make_preview_photo(self, path, max_px=500):
        """Return a Tk-displayable image for ``path``, or None if the format
        cannot be previewed inline. Uses Pillow when available (JPG/BMP/TIFF/PNG/
        GIF/…), falling back to Tk's built-in PNG/GIF support."""
        try:
            from PIL import Image, ImageTk
            with Image.open(path) as im:
                if im.mode not in ("RGB", "RGBA", "L"):
                    im = im.convert("RGBA")
                im.thumbnail((max_px, max_px))
                return ImageTk.PhotoImage(im)
        except ImportError:
            pass  # Pillow not installed — try the Tk-native path below.
        except Exception:
            return None  # Pillow is present but cannot decode this format (e.g. EMF).
        try:
            photo = tk.PhotoImage(file=path)
            w = photo.width()
            if w > max_px:
                f = max(1, w // max_px)
                photo = photo.subsample(f, f)
            return photo
        except Exception:
            return None

    def _on_image_select(self, event=None):
        sel = self.lst_images.curselection()
        if not sel:
            return
        filename = self.lst_images.get(sel[0])
        path = self._resolve_asset(filename, "images")
        self._show_preview_widget("image")
        if not path:
            self.preview_label.configure(image="", text=f"(file not found)\n{filename}")
            self._preview_img = None
            return
        photo = self._make_preview_photo(path)
        if photo is not None:
            self._preview_img = photo  # keep a ref so Tk doesn't GC it
            self.preview_label.configure(image=self._preview_img, text="")
        else:
            self._preview_img = None
            self.preview_label.configure(
                image="",
                text=f"No inline preview for this format.\nDouble-click to open externally:\n{filename}",
            )

    def _on_table_select(self, event=None):
        sel = self.lst_tables.curselection()
        if not sel:
            return
        filename = self.lst_tables.get(sel[0])
        path = self._resolve_asset(filename, "tables")
        if not path:
            self._show_preview_widget("message")
            self.preview_label.configure(image="", text=f"(file not found)\n{filename}")
            return
        self._show_table_preview(path, filename)

    def _show_table_preview(self, path, filename, max_rows=200, max_cols=30):
        """Render the first sheet of an .xlsx table as a grid via openpyxl."""
        try:
            import openpyxl
        except ImportError:
            self._show_preview_widget("message")
            self.preview_label.configure(
                image="",
                text="openpyxl is not installed — cannot preview tables inline.\n"
                     "Double-click to open the file externally.",
            )
            return
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                rows.append(row)
                if i >= max_rows - 1:
                    break
            wb.close()
        except Exception as exc:  # noqa: BLE001
            self._show_preview_widget("message")
            self.preview_label.configure(image="", text=f"Could not read table:\n{exc}")
            return

        ncols = min(max((len(r) for r in rows), default=0), max_cols)
        cols = [f"c{i}" for i in range(ncols)]
        self.table_preview.delete(*self.table_preview.get_children())
        self.table_preview.configure(columns=cols)
        for i, c in enumerate(cols):
            self.table_preview.heading(c, text=f"Col {i + 1}")
            self.table_preview.column(c, width=110, anchor="w", stretch=False)
        for r in rows:
            vals = ["" if v is None else str(v) for v in r][:ncols]
            vals += [""] * (ncols - len(vals))
            self.table_preview.insert("", "end", values=vals)
        self._show_preview_widget("table")
        self.status_var.set(f"Table '{filename}': showing {len(rows)} row(s) (capped at {max_rows}).")

    def _open_selected_asset(self, listbox, kind):
        sel = listbox.curselection()
        if not sel:
            return
        filename = listbox.get(sel[0])
        path = self._resolve_asset(filename, kind)
        if path:
            _open_externally(path)
        else:
            messagebox.showinfo("Not found", f"Could not locate the file on disk:\n{filename}")

    # ------------------------------------------------------------- review -- #
    def _iter_nodes(self, nodes=None, depth=0):
        """Yield (node, depth) for every node in the loaded hierarchy, depth-first."""
        if nodes is None:
            nodes = getattr(self, "_hierarchy", []) or []
        for node in nodes:
            yield node, depth
            yield from self._iter_nodes(node.get("children") or [], depth + 1)

    def _show_summary(self):
        """Pop up a document-level overview: totals and a per-type breakdown."""
        if not getattr(self, "_hierarchy", None):
            messagebox.showinfo("Nothing loaded", "Load a structured JSON first.")
            return
        from collections import Counter

        types = Counter()
        totals = {"nodes": 0, "images": 0, "tables": 0, "hyperlinks": 0, "text": 0}
        for node, _ in self._iter_nodes():
            totals["nodes"] += 1
            types[node.get("element_type", "?")] += 1
            totals["images"] += len(node.get("extracted_images") or [])
            totals["tables"] += len(node.get("extracted_tables") or [])
            totals["hyperlinks"] += len(node.get("hyperlinks") or [])
            totals["text"] += len(node.get("text_content", "") or "")

        meta = self.data.get("document_metadata", {}) if isinstance(self.data, dict) else {}
        stem = Path(self.json_path).stem if self.json_path else ""
        lines = [
            f"Document : {meta.get('source-title', stem)}",
            f"Domain   : {meta.get('Domain', '')}",
            "",
            f"Total nodes      : {totals['nodes']}",
            f"Total hyperlinks : {totals['hyperlinks']}",
            f"Total images     : {totals['images']}",
            f"Total tables     : {totals['tables']}",
            f"Total text       : {totals['text']:,} chars",
            "",
            "Nodes by element type:",
        ]
        for t, c in types.most_common():
            lines.append(f"   {t:<22} {c}")

        win = tk.Toplevel(self.root)
        win.title("Document Summary")
        win.geometry("480x440")
        txt = scrolledtext.ScrolledText(win, wrap="word", font=("TkFixedFont", 10))
        txt.insert("1.0", "\n".join(lines))
        txt.configure(state="disabled")
        txt.pack(fill="both", expand=True, padx=8, pady=8)
        ttk.Button(win, text="Close", command=win.destroy).pack(pady=(0, 8))

    # ------------------------------------------------------------- export -- #
    def _build_export_menu(self, bar):
        mb = ttk.Menubutton(bar, text="Export ▾")
        menu = tk.Menu(mb, tearoff=False)
        menu.add_command(label="Node index → CSV…", command=lambda: self._export_index("csv"))
        menu.add_command(label="Node index → Excel…", command=lambda: self._export_index("xlsx"))
        menu.add_separator()
        menu.add_command(label="All text → Markdown…", command=self._export_text)
        menu.add_command(label="Selected subtree → JSON…", command=self._export_subtree)
        mb["menu"] = menu
        mb.pack(side="right", padx=(0, 6))

    def _default_export_name(self, kind, ext):
        stem = Path(self.json_path).stem if self.json_path else "easa"
        return f"{stem}_{kind}.{ext}"

    def _flatten_rows(self):
        """Flatten the hierarchy into one row per node — an overview table."""
        rows = []
        for node, depth in self._iter_nodes():
            attrs = node.get("attributes", {}) or {}
            text = node.get("text_content", "") or ""
            rows.append({
                "sdt_id": attrs.get("sdt-id", ""),
                "element_type": node.get("element_type", ""),
                "title": attrs.get("source-title") or attrs.get("title") or "",
                "domain": attrs.get("Domain", ""),
                "type_of_content": attrs.get("TypeOfContent", ""),
                "depth": depth,
                "n_children": len(node.get("children") or []),
                "n_hyperlinks": len(node.get("hyperlinks") or []),
                "n_images": len(node.get("extracted_images") or []),
                "n_tables": len(node.get("extracted_tables") or []),
                "text_length": len(text),
                "text_preview": text[:200].replace("\n", " "),
            })
        return rows

    def _export_index(self, fmt):
        if not getattr(self, "_hierarchy", None):
            messagebox.showinfo("Nothing loaded", "Load a structured JSON first.")
            return
        rows = self._flatten_rows()
        headers = list(rows[0].keys()) if rows else []

        if fmt == "csv":
            path = filedialog.asksaveasfilename(
                defaultextension=".csv", filetypes=[("CSV", "*.csv")],
                initialfile=self._default_export_name("node_index", "csv"))
            if not path:
                return
            try:
                import csv
                with open(path, "w", newline="", encoding="utf-8") as f:
                    w = csv.DictWriter(f, fieldnames=headers)
                    w.writeheader()
                    w.writerows(rows)
            except Exception as exc:  # noqa: BLE001
                messagebox.showerror("Export failed", str(exc))
                return
        else:
            try:
                import openpyxl
            except ImportError:
                messagebox.showerror("openpyxl needed", "Install openpyxl to export to Excel.")
                return
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
                initialfile=self._default_export_name("node_index", "xlsx"))
            if not path:
                return
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Node Index"
                ws.append(headers)
                for r in rows:
                    ws.append([r[h] for h in headers])
                wb.save(path)
            except Exception as exc:  # noqa: BLE001
                messagebox.showerror("Export failed", str(exc))
                return

        self.status_var.set(f"Exported {len(rows)} nodes → {path}")
        messagebox.showinfo("Export complete", f"Wrote {len(rows)} rows to:\n{path}")

    def _export_text(self):
        if not getattr(self, "_hierarchy", None):
            messagebox.showinfo("Nothing loaded", "Load a structured JSON first.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".md", filetypes=[("Markdown", "*.md"), ("Text", "*.txt")],
            initialfile=self._default_export_name("text", "md"))
        if not path:
            return
        lines = []
        for node, depth in self._iter_nodes():
            attrs = node.get("attributes", {}) or {}
            title = attrs.get("source-title") or attrs.get("title")
            text = (node.get("text_content", "") or "").strip()
            if title:
                lines.append(f"{'#' * min(depth + 1, 6)} {title}")
            if text:
                lines.append(text)
            if title or text:
                lines.append("")
        try:
            with open(path, "w", encoding="utf-8") as f:
                f.write("\n".join(lines))
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("Export failed", str(exc))
            return
        self.status_var.set(f"Exported document text → {path}")
        messagebox.showinfo("Export complete", f"Wrote text to:\n{path}")

    def _export_subtree(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("No selection", "Select a node in the tree first.")
            return
        node = self.node_by_iid.get(sel[0])
        if node is None:
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".json", filetypes=[("JSON", "*.json")],
            initialfile=self._default_export_name("subtree", "json"))
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(node, f, indent=2, ensure_ascii=False)
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("Export failed", str(exc))
            return
        self.status_var.set(f"Exported selected subtree → {path}")
        messagebox.showinfo("Export complete", f"Wrote selected subtree to:\n{path}")

    # ----------------------------------------------------------------- AI -- #
    def _node_title(self, node):
        a = node.get("attributes", {}) or {}
        return a.get("source-title") or a.get("title") or node.get("element_type", "node")

    def _build_ai_tab(self):
        ai = ttk.Frame(self.main_nb, padding=6)
        self.main_nb.add(ai, text="AI Review")
        self._build_ai_page(ai)

    def _ai_current_section(self):
        if self._current_node is None:
            return None
        node = self._current_node
        return (self._node_title(node), node.get("text_content", "") or "")

    def _ai_checked_sections(self):
        return [(self._node_title(n), n.get("text_content", "") or "")
                for n in self._checked_in_doc_order()]


def main():
    from ..crash_logging import install
    install()
    path = sys.argv[1] if len(sys.argv) > 1 else None
    root = tk.Tk()
    EASAJsonReviewApp(root, json_path=path)
    root.mainloop()


if __name__ == "__main__":
    main()
