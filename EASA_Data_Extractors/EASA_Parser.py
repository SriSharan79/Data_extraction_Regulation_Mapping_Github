import hashlib
import os
import json
import xmltodict
import base64
from pathlib import Path
import re
import zipfile
import io
import xml.etree.ElementTree as ET
import openpyxl  # Ensure you run 'pip3 install openpyxl'

# Master global configurations mapping EASA custom layouts and MS Word OpenXML schemas
NAMESPACES = {
    'pkg': 'http://schemas.microsoft.com/office/2006/xmlPackage',
    'er': 'http://www.easa.europa.eu/erules-export',
    'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main',
    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
    'pic': 'http://schemas.openxmlformats.org/drawingml/2006/picture',
    'wp': 'http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing',
    'v': 'urn:schemas-microsoft-com:vml'
}
def resolve_paths(storage_path, zip_file_path):
    """Generates deterministic target folder paths for the data assets."""
    doc_name = os.path.splitext(os.path.basename(zip_file_path))[0]
    doc_name_2 = doc_name.rsplit('_', 1)[-1]
    target_root = os.path.join(storage_path, doc_name)
    Com_root = os.path.join(storage_path,"All_Combined")
    os.makedirs(Com_root, exist_ok=True)
    Raw_xml_root = os.path.join(Com_root,"Raw_XML_JSONs")
    os.makedirs(Raw_xml_root, exist_ok=True)
    excel_com_root = os.path.join(Com_root,"Overview_excels")
    os.makedirs(excel_com_root, exist_ok=True)
    Structured_Json_root = os.path.join(Com_root,"Structured_JSONs")
    os.makedirs(Structured_Json_root, exist_ok=True)
    base_hash = hashlib.md5(doc_name.encode()).hexdigest()[:8]
    
    return {
        "images_dir": os.path.join(target_root, "images"),
        "tables_dir": os.path.join(target_root, "tables"),
        "output_json": os.path.join(target_root, f"{base_hash}_Extraction_Json.json"),
        "output_json_com": os.path.join(Structured_Json_root, f"{doc_name}.json"),
        "Raw_xml_json_com": os.path.join(Raw_xml_root, f"{doc_name}.json"),
        'master_excel' : os.path.join(target_root, f"Master_Structural_Index.xlsx"),
        'master_excel_com' : os.path.join(excel_com_root,f"{doc_name}.xlsx")
    }
def sanitize_filename(name):
    """Converts a topic title into a safe, clean string usable as an OS filename."""
    if not name:
        return "Unknown_Topic"
    clean = re.sub(r'[^a-zA-Z0-9\s\-\.]', '_', name)
    clean = re.sub(r'[\s_]+', '_', clean)
    return clean.strip('_')[:120]


def build_structural_index(er_doc_node):
    """Builds a flat catalog map of structural nodes based on EASA metadata."""
    flat_nodes = {}
    easa_attributes = [
    'sdt-id',
    'source-title',
    'ERulesId',
    'Domain',
    'ActivityType',
    'AircraftUse',
    'AircraftCategory',
    'AmendedBy',
    'ApplicabilityDate',
    'EntryIntoForceDate',
    'EquivalentForeignRegulation',
    'ICAOReference',
    'Keywords',
    'RegistryState',
    'RegulatedEntity',
    'RegulatorySource',
    'RegulatorySubject',
    'TechnicalSubjectMatter',
    'TypeOfContent',
    'ParentIR',
    'EASACategory',
    'title'
]

    for element in er_doc_node.findall('.//*[@sdt-id]'):
        sdt_id = element.attrib.get('sdt-id')
        tag_name = element.tag.split('}')[-1] if '}' in element.tag else element.tag
        
        attribs = {}
        for attr in easa_attributes:
            if attr in element.attrib and element.attrib[attr].strip():
                attribs[attr] = element.attrib[attr].strip()
                
        raw_name = attribs.get('source-title') or attribs.get('title') or f"{tag_name}_{sdt_id}"
        sanitized_name = sanitize_filename(raw_name)
        
        flat_nodes[sdt_id] = {
            "id": sdt_id,
            "element_type": tag_name,
            "attributes": attribs,
            "sanitized_name": sanitized_name,
            "text_lines": [],
            "hyperlinks": [],
            "extracted_images": [],
            "extracted_tables": [],
            "children_ids": [],
            "image_counter": 0,
            "table_counter": 0
        }

    def map_dependencies(element):
        sdt_id = element.attrib.get('sdt-id')
        for child in element:
            child_sdt_id = child.attrib.get('sdt-id')
            if sdt_id and child_sdt_id and child_sdt_id in flat_nodes:
                # FIX: Mutate global dictionary map instance explicitly
                if child_sdt_id not in flat_nodes[sdt_id]["children_ids"]:
                    flat_nodes[sdt_id]["children_ids"].append(child_sdt_id)
            map_dependencies(child)

    map_dependencies(er_doc_node)
    return flat_nodes


def compile_hierarchy_tree(flat_nodes):
    """Maps individual items into the final recursive JSON tree format."""
    def assemble_nested_tree(node_id):
        raw_data = flat_nodes[node_id]
        
        # FIXED: Ensure all structural fields transfer cleanly to the output JSON blocks
        node_json = {
            "element_type": raw_data["element_type"],
            "attributes": raw_data["attributes"],
            "extracted_images": raw_data["extracted_images"],
            "extracted_tables": raw_data["extracted_tables"],
            "children_ids": raw_data["children_ids"],
            "image_counter": raw_data["image_counter"],
            "table_counter": raw_data["table_counter"]
        }
        
        text_content = "".join(raw_data["text_lines"]).strip()
        if text_content:
            node_json["text_content"] = text_content
        if raw_data["hyperlinks"]:
            node_json["hyperlinks"] = raw_data["hyperlinks"]
            
        if raw_data["children_ids"]:
            node_json["children"] = [assemble_nested_tree(c_id) for c_id in raw_data["children_ids"]]
            
        return node_json

    all_children_ids = set()
    for n in flat_nodes.values():
        all_children_ids.update(n["children_ids"])
        
    root_nodes = [n_id for n_id in flat_nodes if n_id not in all_children_ids]
    return [assemble_nested_tree(r_id) for r_id in root_nodes]


def parse_document_body_stream(root, flat_nodes, rels_map, media_parts, images_dir, tables_dir):
    """Parses paragraphs, tables, and binary shapes out of the layout content stream."""
    print("[DEBUG Body] Locating primary /word/document.xml package entry point...")
    doc_part = root.find(".//pkg:part[@pkg:name='/word/document.xml']", NAMESPACES)
    if doc_part is None:
        print("[CRITICAL] Master OpenXML document text engine target component missing.")
        return

    print("[DEBUG Body] Indexing element ancestry trees for safe upward structural context lookups (This may take a brief moment)...")
    parent_map = {c: p for p in doc_part.findall('.//*') for c in p}
    print(f"[DEBUG Body] Element ancestry tree indexing complete. Total indexed layout tokens: {len(parent_map)}")

    def find_ancestor_sdt_id(element):
        """Walks up the XML tree to resolve the nearest active section container."""
        current = element
        while current in parent_map:
            current = parent_map[current]
            tag_local = current.tag.split('}')[-1] if '}' in current.tag else current.tag
            if tag_local == 'sdt':
                sdt_id_node = current.find('.//w:sdtPr/w:id', NAMESPACES)
                if sdt_id_node is not None:
                    found_id = sdt_id_node.attrib.get(f"{{{NAMESPACES['w']}}}val")
                    if found_id in flat_nodes:
                        return found_id
        return None
    
    def find_subsequent_figure_title(element):
        """Looks ahead in the linear XML block context to harvest a real figure caption."""
        current = element
        while current in parent_map and not current.tag.endswith('p'):
            current = parent_map[current]
            
        if current in parent_map:
            parent_block = parent_map[current]
            siblings = list(parent_block)
            try:
                start_idx = siblings.index(current)
                for i in range(start_idx, min(start_idx + 4, len(siblings))):
                    text_pieces = [t.text for t in siblings[i].findall('.//w:t', NAMESPACES) if t.text]
                    full_text = " ".join(text_pieces).strip()
                    if full_text.lower().startswith("figure"):
                        keep_chars = [c for c in full_text if c.isalnum() or c in (' ', '-', '_')]
                        sanitized_title = "".join(keep_chars).strip().replace(" ", "_")
                        sanitized_title = sanitized_title[:15]
                        if sanitized_title:
                            return sanitized_title
            except ValueError:
                pass
        return None
    
    def _process_and_save_image(element, embed_id, rels_map, media_parts, current_store, images_dir, topic_base_name, fallback_type):
        """
        Processes OpenXML drawing elements, parses binary streams from flat packages,
        and saves image records safely onto disk environments.
        """
        rel_target = rels_map.get(embed_id, "")
        media_keys = []
        if rel_target:
            clean_target = rel_target.lstrip('/')
            media_keys = [
                f"/word/{clean_target}" if not clean_target.startswith("word/") else f"/{clean_target}",
                f"/{clean_target}",
                clean_target,
                clean_target.split('/')[-1]
            ]

        # Extract fallback name strings directly from core attributes
        xml_alt_name = None
        doc_pr = element.find('.//wp:docPr', NAMESPACES)
        nv_pic_pr = element.find('.//pic:cNvPr', NAMESPACES)
        
        if nv_pic_pr is not None and nv_pic_pr.attrib.get('name'):
            xml_alt_name = nv_pic_pr.attrib.get('name')
        elif doc_pr is not None and doc_pr.attrib.get('name'):
            xml_alt_name = doc_pr.attrib.get('name')

        image_bytes = None
        matched_key = None
        
        for key in media_keys:
            if key in media_parts and media_parts[key]:
                image_bytes = media_parts[key]
                matched_key = key
                break

        # Flat package lookup recovery if target paths do not match exactly
        if not image_bytes and xml_alt_name:
            numeric_ids = re.findall(r'\d+', xml_alt_name)
            if numeric_ids:
                target_id = numeric_ids[0]
                target_pattern = re.compile(r'image' + target_id + r'\.(png|jpe?g|gif|emf|wmf)$', re.IGNORECASE)
                
                for real_key in media_parts.keys():
                    if target_pattern.search(real_key) or (f"media/image{target_id}." in real_key.lower()):
                        image_bytes = media_parts[real_key]
                        matched_key = real_key
                        break

        if not image_bytes:
            return

        current_store["image_counter"] = current_store.get("image_counter", 0) + 1
        
        _, native_extension = os.path.splitext(matched_key.lower())
        if not native_extension or len(native_extension) > 5:
            native_extension = ".jpg" if fallback_type != "imagedata" else ".png"

        discovered_figure_title = find_subsequent_figure_title(element)
        
        if discovered_figure_title:
            img_filename = f"{discovered_figure_title}{native_extension}"
        elif xml_alt_name and "picture" not in xml_alt_name.lower():
            clean_alt = "".join([c if c.isalnum() else "_" for c in os.path.splitext(xml_alt_name)[0]]).strip("_")
            img_filename = f"{clean_alt}{native_extension}" if clean_alt else f"img_{current_store['id']}_{embed_id}{native_extension}"
        else:
            img_filename = f"{topic_base_name}_{fallback_type}_{current_store['image_counter']}{native_extension}"

        img_path = os.path.join(images_dir, img_filename)
        
        try:
            with open(img_path, 'wb') as img_f:
                img_f.write(image_bytes)
            print(f" [IMAGE SAVED] -> {img_filename} ({len(image_bytes)} bytes)")
            
            if img_filename not in current_store["extracted_images"]:
                current_store["extracted_images"].append(img_filename)
        except Exception as e:
            print(f" [ERROR Image] Failed processing binary write cycle onto disk path {img_path}: {e}")  
    
    active_node_id = None
    print("[DEBUG Body] Initiating sequential stream loop parsing. Processing elements...")

    # Performance Monitoring Counters
    element_count = 0
    text_count = 0
    table_count = 0

    for element in doc_part.findall('.//*'):
        tag_local = element.tag.split('}')[-1] if '}' in element.tag else element.tag
        element_count += 1
        
        # Log basic checkpoint telemetry output every 25,000 raw layout items parsed
        if element_count % 25000 == 0:
            print(f" [PROGRESS] Parsed {element_count} structural tokens... (Text chunks: {text_count}, Table layouts found: {table_count})")

        if tag_local == 'sdt':
            sdt_id_node = element.find('.//w:sdtPr/w:id', NAMESPACES)
            if sdt_id_node is not None:
                found_id = sdt_id_node.attrib.get(f"{{{NAMESPACES['w']}}}val")
                if found_id in flat_nodes:
                    active_node_id = found_id

        current_context_id = active_node_id if active_node_id else find_ancestor_sdt_id(element)
        if not current_context_id:
            continue
            
        current_store = flat_nodes[current_context_id]
        topic_base_name_raw = current_store["sanitized_name"]
        # print(f'topic_base_name_raw={current_store["sanitized_name"]}')
        topic_base_name = topic_base_name_raw[:10]

        # A. Core Paragraph Text Extraction Logic
        if tag_local == 't' and element.text and element.text.strip():
            text_val = element.text.strip()
            text_count += 1
            if text_val not in current_store["text_lines"]:
                current_store["text_lines"].append(text_val)

        # B. Document Hyperlink Parsing Logic
        elif tag_local == 'hyperlink':
            r_id = element.attrib.get(f"{{{NAMESPACES['r']}}}id")
            link_text = "".join([t.text for t in element.findall('.//w:t', NAMESPACES) if t.text]).strip()
            if link_text:
                target_url = rels_map.get(r_id, "Internal Jump Link")
                current_store["hyperlinks"].append({"text": link_text, "target": target_url})

        # C. Standard OpenXML Document Drawings
        elif tag_local == 'drawing':
            pic_nodes = element.findall('.//pic:pic', NAMESPACES)
            if pic_nodes:
                for pic in pic_nodes:
                    blip = pic.find('.//pic:blipFill/a:blip', NAMESPACES)
                    if blip is None: continue
                    embed_id = blip.attrib.get(f"{{{NAMESPACES['r']}}}embed")
                    if embed_id:
                        _process_and_save_image(element, embed_id, rels_map, media_parts, current_store, images_dir, topic_base_name, "fig")
            else:
                embed_ids = []
                for blip in element.findall('.//a:blip', NAMESPACES):
                    eid = blip.attrib.get(f"{{{NAMESPACES['r']}}}embed")
                    if eid: embed_ids.append(eid)
                    
                for embed_id in embed_ids:
                    _process_and_save_image(element, embed_id, rels_map, media_parts, current_store, images_dir, topic_base_name, "drawing")

        # D. Legacy VML Shapes (imagedata blocks)
        elif tag_local == 'imagedata':
            r_id = element.attrib.get(f"{{{NAMESPACES['r']}}}id")
            if r_id:
                _process_and_save_image(element, r_id, rels_map, media_parts, current_store, images_dir, topic_base_name, "fallback")

        # E. Table Matrix Layout Extraction Engine
        elif tag_local == 'tbl':
            table_count += 1
            if "table_counter" not in current_store:
                current_store["table_counter"] = 0
            current_store["table_counter"] += 1
            
            table_filename = f"{topic_base_name}_{current_store['table_counter']}.xlsx"
            single_table_path = os.path.join(tables_dir, table_filename)
            
            single_wb = openpyxl.Workbook()
            ws = single_wb.active
            ws.title = "Extracted Grid Data"
            
            has_data = False
            for row in element.findall('.//w:tr', NAMESPACES):
                row_cells = []
                for cell in row.findall('.//w:tc', NAMESPACES):
                    cell_str = " ".join([t.text for t in cell.findall('.//w:t', NAMESPACES) if t.text]).strip()
                    row_cells.append(cell_str)
                if any(row_cells):
                    ws.append(row_cells)
                    has_data = True
            
            if has_data:
                single_wb.save(single_table_path)
                print(f" [TABLE SAVED] -> Excel Sheet Matrix: {table_filename}")
                if table_filename not in current_store["extracted_tables"]:
                    current_store["extracted_tables"].append(table_filename)
            else:
                try: single_wb.close()
                except Exception: pass

    print(f"[DEBUG Body] Parsing completed successfully. Total tokens read: {element_count}")
    
def generate_master_excel_index(flat_nodes, output_excel_path):
    """Generates a structured Master Index Excel file tracking all parsed keys, attributes, and text metrics."""
    print(f"[DEBUG Excel] Compiling structural layout tracking matrix to: {output_excel_path}")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Index Catalog"
    
    # 22 Defined explicit EASA attributes + asset paths metadata columns
    easa_attributes = [
        'sdt-id', 'source-title', 'ERulesId', 'Domain', 'ActivityType',
        'AircraftUse', 'AircraftCategory', 'AmendedBy', 'ApplicabilityDate',
        'EntryIntoForceDate', 'EquivalentForeignRegulation', 'ICAOReference',
        'Keywords', 'RegistryState', 'RegulatedEntity', 'RegulatorySource',
        'RegulatorySubject', 'TechnicalSubjectMatter', 'TypeOfContent',
        'ParentIR', 'EASACategory', 'title'
    ]
    
    # Expanded headers list to include requested metrics columns
    headers = (
        ["Node ID", "Element Type", "Sanitized Name File Identifier"] 
        + easa_attributes 
        + ["Length of Text Content", "Number of Hyperlinks","Associated Images Saved", "Associated Tables Saved" ]
    )
    ws.append(headers)
    
    for sdt_id, node in flat_nodes.items():
        attribs = node.get("attributes", {})
        
        row = [
            node.get("id", ""),
            node.get("element_type", ""),
            node.get("sanitized_name", "")
        ]
        
        # Pull each attribute value cleanly
        for attr in easa_attributes:
            row.append(attribs.get(attr, ""))
            
        
        # 1. Calculate Length of Text Content (total character length of all text lines combined)
        combined_text = " ".join(node.get("text_lines", []))
        row.append(len(combined_text))
        
        # 2. Calculate Number of Hyperlinks
        num_hyperlinks = len(node.get("hyperlinks", []))
        row.append(num_hyperlinks)
        
        
        # Append image and table filenames as comma-separated lists
        row.append(", ".join(node.get("extracted_images", [])))
        row.append(", ".join(node.get("extracted_tables", [])))
        
        ws.append(row)
        
    # Apply clean auto-fit margins formatting across table grids
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 11), 50)
        
    wb.save(output_excel_path)
    print(f"[DEBUG Excel] Master Structural Excel Sheet generated successfully.")
    


def extract_easa_from_zip_v3(zip_path, storage_path):
    """Primary orchestration wrapper logic processing workspace initialization and extraction cycles."""
    print(f"\n[DEBUG Main] Opening target archive: {zip_path}")
    
    paths = resolve_paths(storage_path, zip_path)
    os.makedirs(paths["images_dir"], exist_ok=True)
    os.makedirs(paths["tables_dir"], exist_ok=True)
    

    root = None
    xml_data=None
    print("[DEBUG Main] Looking inside ZIP for valid flat XML Packages...")
    with zipfile.ZipFile(zip_path, 'r') as archive:
        for info in archive.infolist():
            if info.filename.lower().endswith('.xml') and not info.filename.startswith('__MACOSX'):
                print(f"[DEBUG Main] Extracting structural stream data from flat item: {info.filename}")
                xml_data = archive.read(info.filename)
                root = ET.fromstring(xml_data)
                break
    
    print("\n--- PHASE 0: RAW COVERSION OF XML TO JSON ---")
     # Convert XML to a Python dictionary
    # xml_attribs=True ensures attributes like "guid" and "pub-time" are kept
    data_dict = xmltodict.parse(xml_data, xml_attribs=True)

    print("Writing JSON file...")
    with open(paths["Raw_xml_json_com"], 'w', encoding='utf-8') as json_file:
        # indent=4 makes the resulting JSON beautifully formatted and readable
        json.dump(data_dict, json_file, indent=4, ensure_ascii=False)

    print(f"Success! Converted JSON saved as: {paths["Raw_xml_json_com"]}")

    if root is None:
        print("[CRITICAL Main] Failed: No XML data elements recovered out of ZIP target root.")
        return

    # Phase 1: Pre-mapping Relationships and Media Content Blocks
    print("\n--- PHASE 1: PRE-MAPPING RELATIONSHIPS & BASE64 BINARIES ---")
    rels_map = {}
    media_parts = {}

    for part in root.findall(".//pkg:part", NAMESPACES):
        part_name = part.attrib.get(f"{{{NAMESPACES['pkg']}}}name", "")
        
        # 1A. Ingest document relationships
        if "document.xml.rels" in part_name or part_name.endswith(".xml.rels"):
            for rel in part.findall(".//Relationships/Relationship", {'': 'http://schemas.openxmlformats.org/package/2006/relationships'}):
                rid = rel.attrib.get("Id")
                target = rel.attrib.get("Target")
                if rid and target:
                    rels_map[rid] = target
            # Support un-namespaced relationship components
            for rel in part.findall(".//Relationship"):
                rid = rel.attrib.get("Id")
                target = rel.attrib.get("Target")
                if rid and target:
                    rels_map[rid] = target

        # 1B. Decode base64 Media Streams natively
        if "/word/media/" in part_name:
            string_data = part.find('.//pkg:stringData', NAMESPACES)
            binary_data = part.find('.//pkg:binaryData', NAMESPACES)
            target_node = string_data if string_data is not None else binary_data
            
            if target_node is not None and target_node.text:
                try:
                    media_parts[part_name] = base64.b64decode(target_node.text.strip())
                except Exception as e:
                    print(f" [WARNING] Failed parsing file binary content mapping stream {part_name}: {e}")

    print(f"[DEBUG Phase 1] Loaded {len(rels_map)} relationship items and {len(media_parts)} binary file components.")

    # Phase 2: Metadata Processing
    print("\n--- PHASE 2: METADATA & STRUCTURAL INDEXING ---")
    er_doc_node = root.find(".//pkg:part/pkg:xmlData/er:document", NAMESPACES)
    if er_doc_node is None:
        print("[CRITICAL Main] Document metadata tree target root reference broken.")
        return

    document_metadata = {
        "guid": er_doc_node.attrib.get('guid', ''),
        "pub-time": er_doc_node.attrib.get('pub-time', ''),
        "source-title": er_doc_node.attrib.get('source-title', ''),
        "Domain": er_doc_node.attrib.get('Domain', '')
    }
    print(f"[DEBUG Phase 2] Source document identified: '{document_metadata['source-title']}'")

    flat_nodes = build_structural_index(er_doc_node)

    # Phase 3: Content Streaming and Media Extraction Work cycles
    print("\n--- PHASE 3: NARRATIVE PARSING & FILE WRITING ---")
    parse_document_body_stream(root, flat_nodes, rels_map, media_parts,  paths["images_dir"], paths["tables_dir"])

    # Phase 4: Hierarchy Construction & JSON Export Configuration
    print("\n--- PHASE 4: COMPILING RECURSIVE OUTPUT TREE ---")
    hierarchy_tree = compile_hierarchy_tree(flat_nodes)
    
    print("\n--- PHASE 4: COMPILING RECURSIVE OUTPUT TREE & MASTER EXCEL ---")
    # Export structural index attributes database directly to table tracking grids (Excludes texts / hyperlinks)
    generate_master_excel_index(flat_nodes,paths["master_excel"] )
    generate_master_excel_index(flat_nodes,paths["master_excel_com"] )

    final_output = {
        "document_metadata": document_metadata,
        "rules_hierarchy": hierarchy_tree
    }

    print(f"Saving compiled extraction data to file path: {paths['output_json']}")
    print(f"Saving compiled extraction data to file path: {paths['output_json_com']}")
    with open(paths["output_json"], 'w', encoding='utf-8') as json_f:
        json.dump(final_output, json_f, indent=4, ensure_ascii=False)
    with open(paths["output_json_com"], 'w', encoding='utf-8') as json_f:
        json.dump(final_output, json_f, indent=4, ensure_ascii=False)
        
    print("\n[SUCCESS] Extraction pipeline has finished running cleanly!")


if __name__ == "__main__":
    # Test script runtime values
    source_path=r"C:\Users\kata_du\Documents\Literature\EASA\Latest_zip files"
    workspace_directory = r"C:\Users\kata_du\Documents\Literature\EASA\XML _Data_extractions\Bulk_Extraction"
    # target_zip =r"C:\Users\kata_du\Documents\Literature\EASA\Latest_zip files\313A4D_2025-11-27_11.38.35_EAR-for-Initial-Airworthiness-and-Environmental-Protection-Regulation-EU-No-748-2012.zip"
    
    # extract_easa_from_zip_v3(target_zip, workspace_directory)
    source_root = Path(source_path)
    
    # # rglob("*.pdf") finds all PDFs in all subfolders
    for file_path in source_root.rglob("*.zip"): 
        target_zip =r"C:\Users\kata_du\Documents\Literature\EASA\Latest_zip files\EAR for CS-25 Amdt 27 (xml) fix 12.22 FINAL (1).zip"

        if os.path.exists(file_path):
            extract_easa_from_zip_v3(file_path, workspace_directory)
        else:
            print(f"Error: Missing verification path file -> {target_zip}")